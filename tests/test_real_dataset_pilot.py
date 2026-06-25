from __future__ import annotations

import json
import io
import unittest
from contextlib import redirect_stdout
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from openpyxl import Workbook

from scripts.generate_domain_pack import inspect_source_dataset
from scripts.run_real_dataset_pilot import (
    _admissions_rows,
    _target_query_record,
    main as pilot_main,
)
from src.api.dataset_service import DatasetService, DatasetServiceError
from tests.workbench_contract_utils import assert_workbench_contract


GROUP_DETAIL_QUERY = (
    "列出25年深圳大学录取最高的专业组及专业组里面的各个专业最低录取分数"
)
RECOMMENDATION_QUERY = (
    "假设我今年的高考分数是630分，位次9000，想读人工智能，计算机，而且不想去国外，"
    "想留在广东省，请给出推荐"
)
ADMISSIONS_HEADERS = [
    "ID",
    "年份",
    "批次",
    "院校代码",
    "院校名称",
    "院校专业组代码",
    "专业组名称",
    "科类",
    "选科要求",
    "专业代码",
    "专业名称",
    "专业全称",
    "所在省",
    "城市",
    "学费",
    "专业组最低位次1",
    "最低位次1",
    "专业组最低分1",
    "最低分1",
    "最高分1",
    "计划人数",
    "专业组计划人数",
    "生源地",
    "院校标签",
    "院校排名",
]


class RealDatasetPilotTest(unittest.TestCase):
    def test_multi_sheet_excel_profile_and_header_detection(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            source = _write_real_like_workbook(root)
            service = DatasetService(root / "managed")

            upload = service.upload(
                filename=source.name,
                content=source.read_bytes(),
                dataset_id="ds_sheet_profile",
            )
            generated = service.generate_domain_pack(
                "ds_sheet_profile",
                domain_name="admissions",
                template_id="admissions_schema_v1",
            )
            profile = service.profile("ds_sheet_profile")
            review = service.review_summary("ds_sheet_profile")

        self.assertEqual(upload["sheet_name"], "招生数据")
        self.assertEqual(upload["detected_header_row"], 3)
        self.assertEqual(profile["detected_header_row"], 3)
        self.assertEqual(generated["profile"]["row_count"], 6)
        summaries = {item["sheet_name"]: item for item in profile["sheet_summaries"]}
        self.assertFalse(summaries["说明"]["selected"])
        self.assertTrue(summaries["招生数据"]["selected"])
        self.assertFalse(review["missing_fields"])

    def test_duplicate_columns_and_empty_rows_columns_cleanup(self) -> None:
        with TemporaryDirectory() as directory:
            source = Path(directory) / "duplicate.csv"
            source.write_text(
                "城市,城市,租金,,空列\n深圳,广州,2000,,\n\n",
                encoding="utf-8",
            )

            inspection = inspect_source_dataset(source)

        self.assertEqual(inspection.dataset.headers, ["城市", "城市_2", "租金"])
        self.assertEqual(len(inspection.dataset.dataframe), 1)
        warning_codes = {item["code"] for item in inspection.warnings}
        self.assertIn("duplicate_columns_normalized", warning_codes)
        self.assertIn("empty_columns_dropped", warning_codes)

    def test_formula_and_merged_cell_warnings_are_structured(self) -> None:
        with TemporaryDirectory() as directory:
            source = _write_real_like_workbook(Path(directory))
            service = DatasetService(Path(directory) / "managed")

            upload = service.upload(
                filename=source.name,
                content=source.read_bytes(),
                dataset_id="ds_structural_warnings",
            )

        warning_codes = {item["code"] for item in upload["warnings"]}
        self.assertIn("merged_cells_detected", warning_codes)
        self.assertIn("hidden_rows_or_columns_detected", warning_codes)
        self.assertIn("formula_cells_detected", warning_codes)

    def test_large_file_guard_returns_structured_error(self) -> None:
        with TemporaryDirectory() as directory:
            service = DatasetService(Path(directory) / "managed")

            with patch("src.api.dataset_service.MAX_UPLOAD_BYTES", 8):
                with self.assertRaises(DatasetServiceError) as context:
                    service.upload(
                        filename="large.csv",
                        content=b"city,rent\nshenzhen,2000\n",
                    )

        self.assertEqual(context.exception.code, "upload_too_large")
        self.assertEqual(context.exception.status_code, 413)
        self.assertIn("max_size_bytes", context.exception.details)

    def test_missing_required_admissions_fields_returns_blocked(self) -> None:
        with TemporaryDirectory() as directory:
            service, dataset_id = _queryable_uploaded_admissions(
                Path(directory),
                rows=[
                    {
                        "年份": 2025,
                        "院校名称": "深圳大学",
                        "院校专业组代码": "10590225",
                        "专业组名称": "物理225组",
                        "专业名称": "人工智能",
                        "专业全称": "人工智能",
                        "专业组最低分1": 634,
                        "专业组最低位次1": 7800,
                    }
                ],
            )

            response = service.query(
                dataset_id,
                user_input=GROUP_DETAIL_QUERY,
                soft_preferences={"prompt": GROUP_DETAIL_QUERY},
                planner_mode="legacy",
            )

        assert_workbench_contract(self, response)
        self.assertEqual(response["status"], "blocked")
        self.assertIn("missing_required_admissions_fields", _warning_codes(response))
        self.assertEqual(response["debug_trace"]["execution"]["sql"], "")

    def test_ambiguous_score_fields_return_blocked_needs_review(self) -> None:
        with TemporaryDirectory() as directory:
            service, dataset_id = _queryable_uploaded_admissions(
                Path(directory),
                rows=[
                    {
                        "年份": 2025,
                        "院校名称": "深圳大学",
                        "院校专业组代码": "10590225",
                        "专业组名称": "物理225组",
                        "专业代码": "080717",
                        "专业名称": "人工智能",
                        "专业全称": "人工智能",
                        "最低分": 632,
                    }
                ],
            )

            response = service.query(
                dataset_id,
                user_input=GROUP_DETAIL_QUERY,
                soft_preferences={"prompt": GROUP_DETAIL_QUERY},
                planner_mode="legacy",
            )

        assert_workbench_contract(self, response)
        self.assertEqual(response["status"], "blocked")
        self.assertIn("ambiguous_admissions_score_fields", _warning_codes(response))
        self.assertEqual(response["debug_trace"]["execution"]["sql"], "")

    def test_uploaded_real_like_fixture_runs_group_detail_report(self) -> None:
        with TemporaryDirectory() as directory:
            service, dataset_id = _queryable_real_like_dataset(Path(directory))

            response = service.query(
                dataset_id,
                user_input=GROUP_DETAIL_QUERY,
                soft_preferences={"prompt": GROUP_DETAIL_QUERY},
                planner_mode="legacy",
            )

        assert_workbench_contract(self, response)
        self.assertEqual(response["status"], "ok")
        self.assertEqual(response["query_type"], "group_detail_report")
        self.assertEqual(response["evidence_pack"]["execution_summary"]["params"][0], 2025)
        self.assertEqual(
            response["evidence_pack"]["execution_summary"]["metric"]["field_id"],
            "group_min_score_2024",
        )
        self.assertTrue(response["result_sections"]["groups"][0]["majors"])

    def test_uploaded_real_like_fixture_runs_recommendation(self) -> None:
        with TemporaryDirectory() as directory:
            service, dataset_id = _queryable_real_like_dataset(Path(directory))

            response = service.query(
                dataset_id,
                user_input=RECOMMENDATION_QUERY,
                soft_preferences={"prompt": RECOMMENDATION_QUERY},
                planner_mode="legacy",
            )

        assert_workbench_contract(self, response)
        self.assertEqual(response["status"], "ok")
        self.assertEqual(response["query_type"], "recommendation")
        self.assertNotIn("score_without_rank", _warning_codes(response))
        self.assertNotIn("录取概率高", response["answer"])
        self.assertTrue(response["items"])

    def test_pilot_report_json_and_markdown_are_generated(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            source = _write_real_like_workbook(root)
            output_dir = root / "pilot"

            with redirect_stdout(io.StringIO()):
                exit_code = pilot_main([str(source), "--output-dir", str(output_dir)])

            report = json.loads((output_dir / "report.json").read_text(encoding="utf-8"))
            markdown = (output_dir / "report.md").read_text(encoding="utf-8")
            serialized = json.dumps(report, ensure_ascii=False)

        self.assertEqual(exit_code, 0)
        self.assertEqual(report["status"], "pass")
        self.assertFalse(report["failures"])
        self.assertIn("target_query_results", report)
        self.assertIn("# Real Dataset Pilot 报告", markdown)
        self.assertNotIn(str(root), serialized)
        self.assertNotIn(str(root), markdown)
        self.assertFalse(Path(str(report["warehouse_path"])).is_absolute())

    def test_pilot_rejects_score_only_recommendation_execution(self) -> None:
        query = (
            "假设我今年的高考分数是630分，想读人工智能，计算机，而且不想去国外，"
            "想留在广东省，请给出推荐"
        )
        record = _target_query_record(
            query,
            {
                "status": "ok",
                "query_type": "recommendation",
                "result_count": 3,
                "items": [{"university_name": "深圳大学"}],
                "warnings": [{"code": "score_without_rank"}],
                "answer": "请补充位次。",
                "evidence_pack": {
                    "execution_summary": {
                        "sql": "SELECT * FROM admissions",
                        "params": [630],
                    }
                },
            },
        )

        self.assertIn(
            "score-only recommendation must return needs_confirmation",
            record["failures"],
        )
        self.assertIn(
            "score-only recommendation must return zero results",
            record["failures"],
        )
        self.assertIn(
            "score-only recommendation must not execute SQL",
            record["failures"],
        )


def _queryable_real_like_dataset(root: Path) -> tuple[DatasetService, str]:
    return _queryable_uploaded_admissions(root, rows=_admissions_rows(), excel=True)


def _queryable_uploaded_admissions(
    root: Path,
    *,
    rows: list[dict[str, object]],
    excel: bool = False,
) -> tuple[DatasetService, str]:
    source = (
        _write_admissions_workbook(root, rows=rows)
        if excel
        else _write_admissions_csv(root, rows=rows)
    )
    dataset_id = "ds_real_like_admissions"
    service = DatasetService(root / "managed")
    service.upload(
        filename=source.name,
        content=source.read_bytes(),
        dataset_id=dataset_id,
    )
    service.generate_domain_pack(
        dataset_id,
        domain_name="admissions",
        template_id="admissions_schema_v1",
    )
    approved = service.approve_domain(dataset_id)
    if not approved["ok"]:
        raise AssertionError(json.dumps(approved, ensure_ascii=False, indent=2))
    built = service.build_warehouse(dataset_id)
    if built["status"] != "queryable":
        raise AssertionError(json.dumps(built, ensure_ascii=False, indent=2))
    return service, dataset_id


def _write_real_like_workbook(root: Path) -> Path:
    return _write_admissions_workbook(root, rows=_admissions_rows(), with_preamble=True)


def _write_admissions_workbook(
    root: Path,
    *,
    rows: list[dict[str, object]],
    with_preamble: bool = False,
) -> Path:
    path = root / "real_like_admissions.xlsx"
    workbook = Workbook()
    workbook.active.title = "说明"
    data = workbook.create_sheet("招生数据")
    if with_preamble:
        data.append(["广东招生数据 pilot fixture"])
        data.append(["系统说明：真实文件常见第一行不是表头"])
    data.append(ADMISSIONS_HEADERS)
    for row in rows:
        data.append([row.get(header) for header in ADMISSIONS_HEADERS])
    if with_preamble:
        data.merge_cells("A1:C1")
        data.row_dimensions[2].hidden = True
        data["Z4"] = "=SUM(U4:V4)"
    workbook.save(path)
    return path


def _write_admissions_csv(root: Path, *, rows: list[dict[str, object]]) -> Path:
    path = root / "admissions.csv"
    headers = list(dict.fromkeys(key for row in rows for key in row.keys()))
    with path.open("w", encoding="utf-8") as handle:
        handle.write(",".join(headers) + "\n")
        for row in rows:
            handle.write(",".join(str(row.get(header, "")) for header in headers) + "\n")
    return path


def _warning_codes(response: dict[str, object]) -> set[str]:
    return {warning["code"] for warning in response["warnings"]}


if __name__ == "__main__":
    unittest.main()
