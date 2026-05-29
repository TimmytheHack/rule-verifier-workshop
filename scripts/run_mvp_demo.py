"""第一个 MVP demo 的最小可执行流水线。

本脚本刻意只服务一个固定输入，不做通用志愿推荐：
1. 只从真实 Excel 表头构建 schema registry。
2. 只执行 deterministic rules 和模拟确认后的 candidate rules。
3. 对缺少字段支撑的“中外合作”偏好明确标记为不执行。
"""

from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


WORKBOOK_NAME = "广东省2025年志愿填报大数据（24-25）0523.xlsx"
OUTPUT_DIR = Path("outputs/mvp_demo")

DEMO_INPUT = "我是广东物理类，排位32000，想学计算机，最好在广州深圳，学校稳一点，不想去太贵的中外合作。"

# MVP 只需要这些列。schema registry 只能由真实存在的列构建，
# 不能因为用户提到了某个概念就发明字段。
REQUIRED_COLUMNS = [
    "生源地",
    "科类",
    "专业名称",
    "城市",
    "专业组最低位次1",
    "学费",
]

DISPLAY_COLUMNS = [
    "ID",
    "年份",
    "批次",
    "院校代码",
    "院校名称",
    "院校专业组代码",
    "专业组名称",
    "专业代码",
    "专业全称",
    "最低位次1",
    "院校标签",
    "院校排名",
]


@dataclass(frozen=True)
class WorkbookContext:
    """保存工作簿中真实表头的位置和列名索引。"""

    sheet_name: str
    header_row: int
    headers: list[str]
    header_index: dict[str, int]


def parse_number(value: Any) -> float | None:
    """把 Excel 中可能是字符串的数字字段解析成 float。"""

    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"\d+(?:\.\d+)?", str(value).replace(",", ""))
    if not match:
        return None
    return float(match.group())


def cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def detect_header_row(ws: Any, required_columns: list[str]) -> WorkbookContext:
    """自动查找包含 MVP 必需列的真实表头行。"""

    required = set(required_columns)
    for row_number, row in enumerate(ws.iter_rows(values_only=True, min_row=1, max_row=25), start=1):
        headers = [cell_text(value) for value in row]
        if required.issubset(set(headers)):
            return WorkbookContext(
                sheet_name=ws.title,
                header_row=row_number,
                headers=headers,
                header_index={name: idx for idx, name in enumerate(headers) if name},
            )
    raise RuntimeError(f"Could not find a header row containing: {', '.join(required_columns)}")


def build_schema_registry(context: WorkbookContext) -> dict[str, dict[str, Any]]:
    """只把真实存在的字段加入 schema registry。

    这是整个 demo 的安全边界：不在 registry 中的字段不能执行。
    """

    candidates = {
        "source_province": {
            "source_column": "生源地",
            "type": "string",
            "aliases": ["广东", "生源地", "考生省份"],
            "allowed_ops": ["eq"],
            "nullable": False,
            "notes": "User source province. For this demo, 广东 is grounded in 生源地.",
        },
        "subject_type": {
            "source_column": "科类",
            "type": "enum",
            "aliases": ["物理类", "历史类", "科类"],
            "allowed_ops": ["eq"],
            "nullable": False,
            "notes": "The demo normalizes 物理类 to 物理.",
        },
        "major_name": {
            "source_column": "专业名称",
            "type": "string",
            "aliases": ["专业", "想学", "计算机"],
            "allowed_ops": ["contains", "eq"],
            "nullable": False,
            "notes": "Only exact keyword matching is deterministic in this MVP.",
        },
        "city": {
            "source_column": "城市",
            "type": "string",
            "aliases": ["城市", "广州", "深圳"],
            "allowed_ops": ["contains", "in_contains"],
            "nullable": True,
            "notes": "The demo uses contains matching for 广州 and 深圳.",
        },
        "group_min_rank_2024": {
            "source_column": "专业组最低位次1",
            "type": "number",
            "aliases": ["专业组最低位次", "2024最低排位", "稳一点"],
            "allowed_ops": [">=", "<=", "between", "sort"],
            "nullable": True,
            "notes": "Used as the confirmed safety-margin field in this MVP.",
        },
        "tuition_yuan_per_year": {
            "source_column": "学费",
            "type": "number_from_string",
            "aliases": ["学费", "太贵", "费用"],
            "allowed_ops": ["<=", ">=", "between"],
            "nullable": True,
            "notes": "Parsed as a number from the Excel cell value.",
        },
    }
    return {
        field_id: spec
        for field_id, spec in candidates.items()
        if spec["source_column"] in context.header_index
    }


def hardcoded_slots() -> dict[str, Any]:
    """第一个 demo 暂时硬编码抽取结果。

    当前研究重点是 rule verification，不是通用中文信息抽取。
    """

    return {
        "input": DEMO_INPUT,
        "user_context": {
            "source_province": "广东",
            "subject_type": "物理",
            "user_rank": 32000,
        },
        "preferences": {
            "major_keyword": "计算机",
            "preferred_cities": ["广州", "深圳"],
            "risk_preference_raw": "稳一点",
            "tuition_preference_raw": "太贵",
            "major_expansion_raw": "计算机相关扩展",
            "cooperation_preference_raw": "不想去太贵的中外合作",
        },
        "raw_phrases": [
            "广东物理类",
            "排位32000",
            "想学计算机",
            "最好在广州深圳",
            "学校稳一点",
            "不想去太贵的中外合作",
        ],
    }


def verify_rule(rule: dict[str, Any], schema_registry: dict[str, dict[str, Any]]) -> dict[str, Any]:
    """验证一条规则是否可以执行。

    Candidate rule 即使字段存在，也因为存在歧义和确认要求而不可直接执行。
    """

    field_id = rule.get("field_id")
    field_exists = field_id in schema_registry
    operator_allowed = field_exists and rule["operator"] in schema_registry[field_id]["allowed_ops"]
    ambiguity_detected = rule["category"] == "candidate"
    requires_confirmation = bool(rule.get("requires_human_confirmation", False))
    executable = field_exists and operator_allowed and not ambiguity_detected and not requires_confirmation
    return {
        "field_exists": field_exists,
        "operator_allowed": operator_allowed,
        "ambiguity_detected": ambiguity_detected,
        "requires_human_confirmation": requires_confirmation,
        "executable": executable,
    }


def build_rules(schema_registry: dict[str, dict[str, Any]]) -> dict[str, Any]:
    """构建 demo 的三类规则和最终可执行规则。"""

    # Deterministic rules：用户表达明确，字段真实存在，操作符被允许。
    deterministic_rules = [
        {
            "rule_id": "d_source_province",
            "source_text": "广东",
            "category": "deterministic",
            "field_id": "source_province",
            "field": "生源地",
            "operator": "eq",
            "value": "广东",
            "confidence": 1.0,
            "requires_human_confirmation": False,
            "trace_reason": "The user explicitly states 广东 and the Excel field 生源地 exists.",
        },
        {
            "rule_id": "d_subject_type",
            "source_text": "物理类",
            "category": "deterministic",
            "field_id": "subject_type",
            "field": "科类",
            "operator": "eq",
            "value": "物理",
            "confidence": 1.0,
            "requires_human_confirmation": False,
            "trace_reason": "The user explicitly states 物理类 and the Excel field 科类 exists.",
        },
        {
            "rule_id": "d_major_keyword",
            "source_text": "想学计算机",
            "category": "deterministic",
            "field_id": "major_name",
            "field": "专业名称",
            "operator": "contains",
            "value": "计算机",
            "confidence": 1.0,
            "requires_human_confirmation": False,
            "trace_reason": "Exact keyword matching is allowed in the MVP.",
        },
        {
            "rule_id": "d_city",
            "source_text": "最好在广州深圳",
            "category": "deterministic",
            "field_id": "city",
            "field": "城市",
            "operator": "in_contains",
            "value": ["广州", "深圳"],
            "confidence": 1.0,
            "requires_human_confirmation": False,
            "trace_reason": "The Excel field 城市 exists and contains relevant values.",
        },
    ]

    # 用户排位是上下文，不是 Excel 过滤字段；它只用于确认后的安全边际计算。
    context_rules = [
        {
            "rule_id": "ctx_user_rank",
            "source_text": "排位32000",
            "category": "context",
            "field_id": "user_rank",
            "field": "user_rank",
            "operator": "set_context",
            "value": 32000,
            "confidence": 1.0,
            "requires_human_confirmation": False,
            "verification": {
                "field_exists": True,
                "operator_allowed": True,
                "ambiguity_detected": False,
                "requires_human_confirmation": False,
                "executable": False,
            },
            "trace_reason": "User rank is context for confirmed candidate rules, not an Excel filter.",
        }
    ]

    # Candidate rules：可以提出候选解释，但在用户确认前绝不能执行。
    candidate_rules = [
        {
            "rule_id": "c_safety_margin",
            "source_text": "学校稳一点",
            "category": "candidate",
            "status": "pending_confirmation",
            "field_id": "group_min_rank_2024",
            "field": "专业组最低位次1",
            "operator": ">=",
            "value_expression_options": ["32000 * 1.05", "32000 * 1.10", "32000 * 1.15"],
            "confidence": 0.8,
            "requires_human_confirmation": True,
            "trace_reason": "稳一点 is vague and must not execute without a confirmed safety margin.",
        },
        {
            "rule_id": "c_tuition_cap",
            "source_text": "太贵",
            "category": "candidate",
            "status": "pending_confirmation",
            "field_id": "tuition_yuan_per_year",
            "field": "学费",
            "operator": "<=",
            "value_options": [10000, 20000, 40000],
            "confidence": 0.8,
            "requires_human_confirmation": True,
            "trace_reason": "太贵 is vague and requires a user-selected tuition cap.",
        },
        {
            "rule_id": "c_major_expansion",
            "source_text": "计算机相关扩展",
            "category": "candidate",
            "status": "pending_confirmation",
            "field_id": "major_name",
            "field": "专业名称",
            "operator": "contains_any",
            "value_options": ["软件工程", "人工智能", "数据科学", "网络空间安全"],
            "confidence": 0.7,
            "requires_human_confirmation": True,
            "trace_reason": "Semantic expansion beyond exact 计算机 matching requires confirmation.",
        },
    ]

    # 中外合作在当前 Excel 中没有专门字段。MVP 不从文本字段里猜测，
    # 因为这会把未验证的语义判断伪装成 deterministic rule。
    llm_needed_parts = [
        {
            "part_id": "l_cooperation_type",
            "source_text": "不想去太贵的中外合作",
            "category": "llm_needed",
            "status": "not_executable_in_mvp",
            "field_id": "cooperation_type",
            "field": "cooperation_type",
            "reason": "The schema registry has no dedicated cooperation_type field.",
            "allowed_behavior": "Do not infer cooperation_type from text fields in this MVP.",
            "verification": {
                "field_exists": "cooperation_type" in schema_registry,
                "schema_grounded": False,
                "executable": False,
            },
            "trace_reason": "No dedicated cooperation_type field, so the preference is preserved but not executed.",
        }
    ]

    # 对 deterministic 和 candidate 统一记录 verification 结果，方便报告审计。
    for rule in deterministic_rules:
        rule["status"] = "verified" if verify_rule(rule, schema_registry)["executable"] else "blocked"
        rule["verification"] = verify_rule(rule, schema_registry)
    for rule in candidate_rules:
        rule["verification"] = verify_rule(rule, schema_registry)

    # 第一个 demo 使用固定的“模拟确认”，避免先做交互 UI。
    simulated_confirmation = {
        "safety_margin": {
            "selected_option": "B",
            "label": "适中稳妥",
            "field": "专业组最低位次1",
            "operator": ">=",
            "value": 35200,
            "source_expression": "32000 * 1.10",
        },
        "tuition_threshold": {
            "selected_option": "B",
            "label": "<= 20000 元/年",
            "field": "学费",
            "operator": "<=",
            "value": 20000,
        },
        "major_expansion": {
            "selected_option": "A",
            "label": "不扩展，只匹配“计算机”",
            "expanded_terms": [],
        },
        "cooperation_type": {
            "selected_option": None,
            "status": "not_executable",
            "reason": "Missing dedicated cooperation_type field.",
        },
    }

    # 最终执行集合 = verified deterministic rules + confirmed candidate rules。
    executable_rules = [
        {
            "rule_id": "e_source_province",
            "derived_from": "d_source_province",
            "field": "生源地",
            "operator": "eq",
            "value": "广东",
        },
        {
            "rule_id": "e_subject_type",
            "derived_from": "d_subject_type",
            "field": "科类",
            "operator": "eq",
            "value": "物理",
        },
        {
            "rule_id": "e_major_keyword",
            "derived_from": "d_major_keyword",
            "field": "专业名称",
            "operator": "contains",
            "value": "计算机",
        },
        {
            "rule_id": "e_city",
            "derived_from": "d_city",
            "field": "城市",
            "operator": "in_contains",
            "value": ["广州", "深圳"],
        },
        {
            "rule_id": "e_safety_margin",
            "derived_from": "c_safety_margin",
            "field": "专业组最低位次1",
            "operator": ">=",
            "value": 35200,
            "confirmation": "safety margin = 10%",
        },
        {
            "rule_id": "e_tuition_cap",
            "derived_from": "c_tuition_cap",
            "field": "学费",
            "operator": "<=",
            "value": 20000,
            "normalization": "parse numeric value from cell text",
            "confirmation": "tuition cap = 20000",
        },
    ]

    return {
        "input": DEMO_INPUT,
        "extracted_slots": hardcoded_slots(),
        "schema_registry": schema_registry,
        "deterministic_rules": deterministic_rules,
        "context_rules": context_rules,
        "candidate_rules": candidate_rules,
        "llm_needed_parts": llm_needed_parts,
        "confirmation_questions": [
            {
                "question_id": "q_safety_margin",
                "source_text": "学校稳一点",
                "options": [
                    {"label": "轻微稳妥", "value": 33600, "expression": "32000 * 1.05"},
                    {"label": "适中稳妥", "value": 35200, "expression": "32000 * 1.10"},
                    {"label": "保守稳妥", "value": 36800, "expression": "32000 * 1.15"},
                    {"label": "不使用这个规则", "value": None},
                ],
            },
            {
                "question_id": "q_tuition_cap",
                "source_text": "太贵",
                "options": [
                    {"label": "<= 10000 元/年", "value": 10000},
                    {"label": "<= 20000 元/年", "value": 20000},
                    {"label": "<= 40000 元/年", "value": 40000},
                    {"label": "不使用学费规则", "value": None},
                ],
            },
            {
                "question_id": "q_major_expansion",
                "source_text": "想学计算机",
                "options": [
                    {"label": "不扩展，只匹配“计算机”", "value": []},
                    {"label": "扩展到相关专业", "value": ["软件工程", "人工智能", "数据科学", "网络空间安全"]},
                ],
            },
        ],
        "simulated_confirmations": simulated_confirmation,
        "final_executable_rules": executable_rules,
        "non_executable_preferences": [
            {
                "source_text": "不想去太贵的中外合作",
                "status": "not_executed",
                "reason": "Missing dedicated cooperation_type field. No text-field inference is used in this MVP.",
            }
        ],
    }


def row_value(row: tuple[Any, ...], context: WorkbookContext, column: str) -> Any:
    """按真实表头索引读取单元格值。"""

    idx = context.header_index[column]
    return row[idx] if idx < len(row) else None


def execute_query(ws: Any, context: WorkbookContext) -> list[dict[str, Any]]:
    """执行最终规则并生成逐行 trace。

    这里使用 AND 逻辑。注意：不包含任何 cooperation_type 过滤，
    因为该字段没有通过 schema grounding。
    """

    results: list[dict[str, Any]] = []
    for excel_row_number, row in enumerate(
        ws.iter_rows(values_only=True, min_row=context.header_row + 1), start=context.header_row + 1
    ):
        if not any(value is not None for value in row):
            continue

        source_province = cell_text(row_value(row, context, "生源地"))
        subject_type = cell_text(row_value(row, context, "科类"))
        major_name = cell_text(row_value(row, context, "专业名称"))
        city = cell_text(row_value(row, context, "城市"))
        group_rank = parse_number(row_value(row, context, "专业组最低位次1"))
        tuition = parse_number(row_value(row, context, "学费"))

        # 六条最终可执行规则。前四条来自 deterministic rules，
        # 后两条来自模拟确认后的 candidate rules。
        checks = [
            source_province == "广东",
            subject_type == "物理",
            "计算机" in major_name,
            any(target_city in city for target_city in ["广州", "深圳"]),
            group_rank is not None and group_rank >= 35200,
            tuition is not None and tuition <= 20000,
        ]
        if not all(checks):
            continue

        # 输出行保留 ranking_key 和安全边际，便于审计排序逻辑。
        output_row = {
            "excel_row_number": excel_row_number,
            "ID": row_value(row, context, "ID") if "ID" in context.header_index else None,
            "年份": row_value(row, context, "年份") if "年份" in context.header_index else None,
            "批次": row_value(row, context, "批次") if "批次" in context.header_index else None,
            "院校代码": row_value(row, context, "院校代码") if "院校代码" in context.header_index else None,
            "院校名称": row_value(row, context, "院校名称") if "院校名称" in context.header_index else None,
            "院校专业组代码": row_value(row, context, "院校专业组代码") if "院校专业组代码" in context.header_index else None,
            "专业组名称": row_value(row, context, "专业组名称") if "专业组名称" in context.header_index else None,
            "专业代码": row_value(row, context, "专业代码") if "专业代码" in context.header_index else None,
            "专业名称": major_name,
            "专业全称": row_value(row, context, "专业全称") if "专业全称" in context.header_index else None,
            "城市": city,
            "学费": tuition,
            "专业组最低位次1": int(group_rank),
            "最低位次1": row_value(row, context, "最低位次1") if "最低位次1" in context.header_index else None,
            "院校标签": row_value(row, context, "院校标签") if "院校标签" in context.header_index else None,
            "院校排名": row_value(row, context, "院校排名") if "院校排名" in context.header_index else None,
            "ranking_key": int(group_rank - 35200),
            "safety_margin_pct": round((group_rank - 32000) / 32000, 4),
            "cooperation_filter_status": "not_executed_missing_cooperation_type_field",
        }
        # trace 必须同时说明“通过了哪些规则”和“哪些用户偏好没有执行”。
        output_row["trace"] = [
            {"rule_id": "e_source_province", "status": "pass", "reason": "生源地 == 广东"},
            {"rule_id": "e_subject_type", "status": "pass", "reason": "科类 == 物理"},
            {"rule_id": "e_major_keyword", "status": "pass", "reason": "专业名称 contains 计算机"},
            {"rule_id": "e_city", "status": "pass", "reason": f"城市 matches {city}"},
            {
                "rule_id": "e_safety_margin",
                "status": "pass",
                "reason": f"专业组最低位次1 {int(group_rank)} >= 35200",
            },
            {"rule_id": "e_tuition_cap", "status": "pass", "reason": f"学费 {tuition:g} <= 20000"},
            {
                "rule_id": "l_cooperation_type",
                "status": "not_executed",
                "reason": "Missing dedicated cooperation_type field; no text inference applied.",
            },
        ]
        results.append(output_row)

    def school_rank_sort(value: Any) -> float:
        # 院校排名有时是 "/" 等非数字值，排序时放到后面。
        parsed = parse_number(value)
        return parsed if parsed is not None else 999999.0

    return sorted(
        results,
        key=lambda item: (
            item["ranking_key"],
            school_rank_sort(item["院校排名"]),
            item["ID"] if isinstance(item["ID"], (int, float)) else 999999,
        ),
    )


def data_coverage(ws: Any, context: WorkbookContext, columns: list[str]) -> dict[str, Any]:
    """统计 MVP 必需字段的数据覆盖率，用于验证报告。"""

    total = 0
    non_null = {column: 0 for column in columns}
    for row in ws.iter_rows(values_only=True, min_row=context.header_row + 1):
        if not any(value is not None for value in row):
            continue
        total += 1
        for column in columns:
            if column in context.header_index and row_value(row, context, column) is not None:
                non_null[column] += 1
    return {
        column: {
            "non_null": non_null[column],
            "total_rows": total,
            "coverage": round(non_null[column] / total, 4) if total else 0,
        }
        for column in columns
    }


def write_rules_json(rules_payload: dict[str, Any]) -> None:
    """输出规则、schema registry、确认结果和最终执行规则。"""

    (OUTPUT_DIR / "rules.json").write_text(
        json.dumps(rules_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def write_filtered_results_csv(results: list[dict[str, Any]]) -> None:
    """输出筛选结果表，供人工检查和后续评估使用。"""

    columns = [
        "rank",
        "excel_row_number",
        "ID",
        "年份",
        "批次",
        "院校代码",
        "院校名称",
        "院校专业组代码",
        "专业组名称",
        "专业代码",
        "专业名称",
        "专业全称",
        "城市",
        "学费",
        "专业组最低位次1",
        "最低位次1",
        "院校标签",
        "院校排名",
        "ranking_key",
        "safety_margin_pct",
        "cooperation_filter_status",
    ]
    with (OUTPUT_DIR / "filtered_results.csv").open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for rank, row in enumerate(results, start=1):
            csv_row = {column: row.get(column) for column in columns}
            csv_row["rank"] = rank
            writer.writerow(csv_row)


def write_verification_report(
    context: WorkbookContext,
    rules_payload: dict[str, Any],
    coverage: dict[str, Any],
    result_count: int,
) -> None:
    """输出可读验证报告，说明哪些规则执行、哪些规则被阻止。"""

    deterministic_lines = []
    for rule in rules_payload["deterministic_rules"]:
        verification = rule["verification"]
        deterministic_lines.append(
            f"| `{rule['rule_id']}` | `{rule['field']}` | `{rule['operator']}` | "
            f"`{rule['value']}` | {verification['executable']} | {rule['trace_reason']} |"
        )

    candidate_lines = []
    for rule in rules_payload["candidate_rules"]:
        candidate_lines.append(
            f"| `{rule['rule_id']}` | {rule['source_text']} | {rule['status']} | "
            f"{rule['requires_human_confirmation']} | {rule['trace_reason']} |"
        )

    coverage_lines = []
    for column, stats in coverage.items():
        coverage_lines.append(
            f"| `{column}` | {stats['non_null']} | {stats['total_rows']} | {stats['coverage']:.2%} |"
        )

    content = f"""# MVP Demo Verification Report

## Input

```text
{DEMO_INPUT}
```

## Workbook

- Sheet: `{context.sheet_name}`
- Detected header row: `{context.header_row}`
- Required demo columns found: `{', '.join(REQUIRED_COLUMNS)}`

## Schema Boundary

The schema registry was built only from real Excel fields needed for this demo.

Missing field:

- `cooperation_type`: not present. The 中外合作 preference is not executable in this MVP.

## Data Coverage

| Column | Non-null | Total rows | Coverage |
|---|---:|---:|---:|
{chr(10).join(coverage_lines)}

## Deterministic Rule Verification

| Rule | Field | Operator | Value | Executable | Reason |
|---|---|---|---|---:|---|
{chr(10).join(deterministic_lines)}

## Candidate Rules

| Rule | Source text | Status | Requires confirmation | Reason |
|---|---|---|---:|---|
{chr(10).join(candidate_lines)}

## Simulated Confirmations

- Safety margin: 10%, so `专业组最低位次1 >= 35200`.
- Tuition cap: `学费 <= 20000`.
- Major expansion: false; only exact keyword `计算机` is used.
- Cooperation type exclusion: not executed because `cooperation_type` is missing.

## Final Query Behavior

The query applies six executable rules with AND logic:

1. `生源地 == 广东`
2. `科类 == 物理`
3. `专业名称 contains 计算机`
4. `城市 contains 广州 or 深圳`
5. `专业组最低位次1 >= 35200`
6. `学费 <= 20000`

The query does not infer or filter 中外合作 from text fields.

## Result Summary

- Filtered row count: `{result_count}`
- Ranking: closest safe professional-group rank first, using `专业组最低位次1 - 35200`.

## Safety Checks

- No LLM is used in code.
- No semantic major expansion is applied.
- No `cooperation_type` field is invented.
- Candidate rules are promoted only through simulated confirmation.
"""
    (OUTPUT_DIR / "verification_report.md").write_text(content, encoding="utf-8")


def write_result_trace(results: list[dict[str, Any]]) -> None:
    """输出逐行 trace，保证每条结果都能追溯到规则。"""

    lines = [
        "# MVP Demo Result Trace",
        "",
        "Every returned row passed the six executable rules. The 中外合作 preference is shown as not executed because the schema lacks a dedicated `cooperation_type` field.",
        "",
        f"Total returned rows: {len(results)}",
        "",
    ]
    for rank, row in enumerate(results, start=1):
        lines.extend(
            [
                f"## {rank}. {row['院校名称']} - {row['专业名称']}",
                "",
                f"- ID: `{row['ID']}`",
                f"- Excel row: `{row['excel_row_number']}`",
                f"- 专业组: `{row['院校专业组代码']} {row['专业组名称']}`",
                f"- 城市: `{row['城市']}`",
                f"- 学费: `{row['学费']:g}`",
                f"- 专业组最低位次1: `{row['专业组最低位次1']}`",
                f"- Ranking key: `{row['ranking_key']}`",
                f"- Safety margin vs user rank: `{row['safety_margin_pct']:.2%}`",
                "",
                "| Rule | Status | Reason |",
                "|---|---|---|",
            ]
        )
        for trace_item in row["trace"]:
            lines.append(
                f"| `{trace_item['rule_id']}` | {trace_item['status']} | {trace_item['reason']} |"
            )
        lines.append("")

    (OUTPUT_DIR / "result_trace.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    """运行第一个 MVP demo。"""

    workbook_path = Path(WORKBOOK_NAME)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    context = detect_header_row(ws, REQUIRED_COLUMNS)
    schema_registry = build_schema_registry(context)

    # 必需列缺失时直接失败，而不是降级为猜测或 LLM 判断。
    missing_required = [column for column in REQUIRED_COLUMNS if column not in context.header_index]
    if missing_required:
        raise RuntimeError(f"Missing required columns: {', '.join(missing_required)}")

    rules_payload = build_rules(schema_registry)
    results = execute_query(ws, context)
    coverage = data_coverage(ws, context, REQUIRED_COLUMNS)

    write_rules_json(rules_payload)
    write_filtered_results_csv(results)
    write_verification_report(context, rules_payload, coverage, len(results))
    write_result_trace(results)

    print(f"Wrote {OUTPUT_DIR / 'rules.json'}")
    print(f"Wrote {OUTPUT_DIR / 'verification_report.md'}")
    print(f"Wrote {OUTPUT_DIR / 'filtered_results.csv'}")
    print(f"Wrote {OUTPUT_DIR / 'result_trace.md'}")
    print(f"Filtered rows: {len(results)}")


if __name__ == "__main__":
    main()
