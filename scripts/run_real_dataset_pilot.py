"""真实上传数据集 pilot：profile、review、建仓和 admissions 查询验收。"""

from __future__ import annotations

import argparse
import json
import math
import shutil
import sys
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from src.api.dataset_service import DatasetService, DatasetServiceError


OUTPUT_DIR = Path("outputs/real_dataset_pilot")
REPORT_JSON_PATH = OUTPUT_DIR / "report.json"
REPORT_MD_PATH = OUTPUT_DIR / "report.md"
UPLOAD_ROOT = OUTPUT_DIR / "uploaded_datasets"
FIXTURE_DIR = OUTPUT_DIR / "fixtures"
TARGET_QUERIES = [
    "列出25年深圳大学录取最高的专业组及专业组里面的各个专业最低录取分数",
    (
        "假设我今年的高考分数是630分，想读人工智能，计算机，而且不想去国外，"
        "想留在广东省，请给出推荐"
    ),
]
REPORT_PATH_KEYS = {
    "source_path",
    "warehouse_path",
    "database_path",
    "domain_dir",
    "schema_profile_path",
    "schema_value_index_path",
    "ingestion_summary_path",
    "json_report",
    "markdown_report",
    "workbook_path",
    "fixture_path",
}


def main(argv: list[str] | None = None) -> int:
    args = _arg_parser().parse_args(argv)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if not args.fixture and not args.source_path:
        raise SystemExit("source_path is required unless --fixture is used.")
    source_path = _fixture_path() if args.fixture else Path(args.source_path)
    report = run_pilot(
        source_path=source_path,
        sheet_name=args.sheet_name,
        output_dir=Path(args.output_dir),
    )
    report_path = Path(args.output_dir)
    report_path.mkdir(parents=True, exist_ok=True)
    (report_path / "report.json").write_text(
        json.dumps(_json_ready(report), ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (report_path / "report.md").write_text(
        _render_markdown(report),
        encoding="utf-8",
    )
    print(f"Wrote {report_path / 'report.md'}")
    print(f"Wrote {report_path / 'report.json'}")
    return 0 if not report["failures"] else 1


def run_pilot(
    *,
    source_path: Path,
    sheet_name: str | None = None,
    output_dir: Path = OUTPUT_DIR,
) -> dict[str, Any]:
    """执行真实数据集 pilot 全链路。"""

    dataset_id = _dataset_id(source_path)
    upload_root = output_dir / "uploaded_datasets"
    shutil.rmtree(upload_root / dataset_id, ignore_errors=True)
    service = DatasetService(upload_root)
    failures: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    target_results: list[dict[str, Any]] = []
    try:
        upload = service.upload(
            filename=source_path.name,
            content=source_path.read_bytes(),
            dataset_id=dataset_id,
            sheet_name=sheet_name,
        )
        warnings.extend(upload.get("warnings") or [])
        generated = service.generate_domain_pack(
            dataset_id,
            domain_name="admissions",
            base_domain="admissions",
        )
        profile = service.profile(dataset_id)
        review = service.review_summary(dataset_id)
        suggestions = _safe_auto_suggest_approvals(review)
        approved = service.approve_domain(dataset_id)
        if not approved.get("ok"):
            failures.append(
                {
                    "stage": "approve-domain",
                    "message": approved.get("message"),
                    "payload": approved.get("payload"),
                }
            )
        warehouse = service.build_warehouse(dataset_id)
        warnings.extend((warehouse.get("warehouse_audit") or {}).get("warnings") or [])
        for query in TARGET_QUERIES:
            response = service.query(
                dataset_id,
                user_input=query,
                soft_preferences={"prompt": query},
                extractor="regex",
                planner_mode="legacy",
            )
            target_results.append(_target_query_record(query, response))
        for result in target_results:
            for failure in result.get("failures") or []:
                failures.append(
                    {
                        "stage": "target-query",
                        "query": result.get("query"),
                        "message": failure,
                    }
                )
    except (DatasetServiceError, FileNotFoundError, ValueError) as exc:
        failures.append(
            {
                "stage": "pilot",
                "message": str(exc),
                "error_type": type(exc).__name__,
            }
        )
        upload = {}
        generated = {}
        profile = {}
        review = {}
        suggestions = []
        approved = {}
        warehouse = {}

    review_payload = (approved.get("payload") or {}).get("review") or {}
    approved_fields = sorted((review_payload.get("reviewed_fields") or {}).keys())
    blocked_fields = sorted((review_payload.get("blocked_fields") or {}).keys())
    warehouse_audit = warehouse.get("warehouse_audit") or {}
    duckdb = warehouse_audit.get("duckdb") or {}
    report_status = "fail" if failures else "pass"
    report = {
        "status": report_status,
        "generated_at": _utc_now(),
        "source_path": str(source_path),
        "dataset_id": upload.get("dataset_id") or dataset_id,
        "source_fingerprint": upload.get("source_fingerprint"),
        "sheet_name": upload.get("sheet_name"),
        "row_count": upload.get("row_count"),
        "column_count": upload.get("column_count"),
        "detected_header_row": upload.get("detected_header_row"),
        "schema_profile_summary": _schema_profile_summary(profile),
        "risky_fields": review.get("risky_fields", []),
        "required_fields": review.get("required_fields", []),
        "missing_fields": review.get("missing_fields", []),
        "safe_auto_suggest_approvals": suggestions,
        "manual_approval_fixture": _manual_approval_fixture(),
        "approved_fields": approved_fields,
        "blocked_fields": blocked_fields,
        "warehouse_path": (warehouse.get("warehouse") or {}).get("database_path"),
        "warehouse_fingerprint": duckdb.get("fingerprint"),
        "target_query_results": target_results,
        "warnings": warnings,
        "failures": failures,
        "artifacts": {
            "domain_dir": generated.get("domain_dir") or upload.get("domain_dir"),
            "schema_profile_path": generated.get("schema_profile_path"),
            "schema_value_index_path": generated.get("schema_value_index_path"),
            "ingestion_summary_path": generated.get("ingestion_summary_path"),
        },
    }
    if any(result["status"] in {"blocked", "error"} for result in target_results):
        report["warnings"].append(
            {
                "code": "target_query_not_fully_executable",
                "severity": "warning",
                "message": "至少一条目标 query 返回 blocked/error，请查看 EvidencePack。",
            }
        )
    return _sanitize_report_paths(report, base_dir=output_dir)


def _target_query_record(query: str, response: dict[str, Any]) -> dict[str, Any]:
    evidence = response.get("evidence_pack") or {}
    execution = evidence.get("execution_summary") or {}
    failures = []
    if response.get("status") == "error":
        failures.append("query returned error")
    if response.get("query_type") == "recommendation":
        warning_codes = [item.get("code") for item in response.get("warnings") or []]
        if "score_without_rank" not in warning_codes:
            failures.append("recommendation missing score_without_rank warning")
        if _is_score_only_recommendation(query, warning_codes):
            if response.get("status") != "needs_confirmation":
                failures.append("score-only recommendation must return needs_confirmation")
            if response.get("result_count") != 0:
                failures.append("score-only recommendation must return zero results")
            if execution.get("sql"):
                failures.append("score-only recommendation must not execute SQL")
            if execution.get("params"):
                failures.append("score-only recommendation must not include SQL params")
        if "录取概率" in str(response.get("answer")) and "不是录取概率" not in str(
            response.get("answer")
        ):
            failures.append("recommendation answer may imply admission probability")
    return {
        "query": query,
        "status": response.get("status"),
        "query_type": response.get("query_type"),
        "result_count": response.get("result_count"),
        "items": response.get("items", []),
        "top_results": response.get("top_results", []),
        "result_sections": response.get("result_sections", {}),
        "warnings": response.get("warnings", []),
        "evidence_pack": evidence,
        "sql": execution.get("sql", ""),
        "params": execution.get("params", []),
        "failures": failures,
    }


def _is_score_only_recommendation(query: str, warning_codes: list[Any]) -> bool:
    if "score_without_rank" in warning_codes:
        return True
    has_score = "分" in query or "成绩" in query
    has_rank = any(term in query for term in ["位次", "排位", "排名", "省排", "省排名"])
    return has_score and not has_rank


def _safe_auto_suggest_approvals(review: dict[str, Any]) -> list[dict[str, Any]]:
    suggestions = []
    for field in review.get("reviewable_fields") or []:
        if field.get("risk_flags"):
            continue
        if not field.get("reviewed"):
            continue
        suggestions.append(
            {
                "field_id": field.get("field_id"),
                "source_column": field.get("source_column"),
                "approved_ops": field.get("approved_ops") or field.get("allowed_ops"),
                "reason": "字段来自已审查 admissions template，仍需 manual approval fixture 落盘。",
            }
        )
    return suggestions


def _manual_approval_fixture() -> dict[str, Any]:
    return {
        "base_domain": "admissions",
        "title_field": "university_name",
        "primary_fields": ["group_code", "major_name", "city"],
        "default_safe_sort": True,
        "note": "pilot 只批准已审查 admissions template，不把自动 seed ops 直接变成 hard rules。",
    }


def _schema_profile_summary(profile: dict[str, Any]) -> dict[str, Any]:
    fields = profile.get("fields") or []
    inferred = Counter(field.get("inferred_type") for field in fields)
    return {
        "row_count": profile.get("row_count"),
        "column_count": profile.get("column_count"),
        "sheet_summaries": profile.get("sheet_summaries", []),
        "detected_header_row": profile.get("detected_header_row"),
        "header_detection_status": profile.get("header_detection_status"),
        "inferred_type_distribution": dict(inferred),
        "warnings": profile.get("warnings", []),
    }


def _fixture_path() -> Path:
    FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
    path = FIXTURE_DIR / "real_like_admissions_pilot.xlsx"
    workbook = Workbook()
    intro = workbook.active
    intro.title = "说明"
    data = workbook.create_sheet("招生数据")
    headers = [
        "ID",
        "年份",
        "批次",
        "院校代码",
        "院校名称",
        "院校专业组代码",
        "专业组名称",
        "科类",
        "选科要求",
        "专业代码",
        "专业名称",
        "专业全称",
        "所在省",
        "城市",
        "学费",
        "专业组最低位次1",
        "最低位次1",
        "专业组最低分1",
        "最低分1",
        "最高分1",
        "计划人数",
        "专业组计划人数",
        "生源地",
        "院校标签",
        "院校排名",
    ]
    data.append(["广东招生数据 pilot fixture"])
    data.append(["系统说明：真实文件常见第一行不是表头"])
    data.append(headers)
    for row in _admissions_rows():
        data.append([row.get(header) for header in headers])
    data.merge_cells("A1:C1")
    data.row_dimensions[2].hidden = True
    data["Z4"] = "=SUM(U4:V4)"
    workbook.save(path)
    return path


def _admissions_rows() -> list[dict[str, Any]]:
    base = {
        "批次": "本科",
        "科类": "物理",
        "选科要求": "化学",
        "学费": 6850,
        "生源地": "广东",
        "院校标签": "省内",
    }
    rows = [
        {
            **base,
            "ID": 1,
            "年份": 2024,
            "院校代码": "10590",
            "院校名称": "深圳大学",
            "院校专业组代码": "10590221",
            "专业组名称": "物理221组",
            "专业代码": "080901",
            "专业名称": "计算机科学与技术",
            "专业全称": "计算机科学与技术",
            "所在省": "广东",
            "城市": "深圳",
            "专业组最低位次1": 9000,
            "最低位次1": 8800,
            "专业组最低分1": 628,
            "最低分1": 626,
            "最高分1": 640,
            "计划人数": 30,
            "专业组计划人数": 100,
            "院校排名": 80,
        },
        {
            **base,
            "ID": 2,
            "年份": 2024,
            "院校代码": "10590",
            "院校名称": "深圳大学",
            "院校专业组代码": "10590221",
            "专业组名称": "物理221组",
            "专业代码": "080717",
            "专业名称": "人工智能",
            "专业全称": "人工智能",
            "所在省": "广东",
            "城市": "深圳",
            "专业组最低位次1": 9000,
            "最低位次1": 9100,
            "专业组最低分1": 628,
            "最低分1": 625,
            "最高分1": 638,
            "计划人数": 20,
            "专业组计划人数": 100,
            "院校排名": 80,
        },
        {
            **base,
            "ID": 3,
            "年份": 2024,
            "院校代码": "10558",
            "院校名称": "中山大学",
            "院校专业组代码": "10558219",
            "专业组名称": "物理219组",
            "专业代码": "080901",
            "专业名称": "计算机类",
            "专业全称": "计算机类",
            "所在省": "广东",
            "城市": "广州",
            "专业组最低位次1": 5000,
            "最低位次1": 5000,
            "专业组最低分1": 650,
            "最低分1": 650,
            "最高分1": 670,
            "计划人数": 40,
            "专业组计划人数": 120,
            "院校排名": 10,
        },
        {
            **base,
            "ID": 4,
            "年份": 2024,
            "院校代码": "11845",
            "院校名称": "广东工业大学",
            "院校专业组代码": "11845101",
            "专业组名称": "物理101组",
            "专业代码": "080717",
            "专业名称": "人工智能",
            "专业全称": "人工智能",
            "所在省": "广东",
            "城市": "广州",
            "专业组最低位次1": 30000,
            "最低位次1": 30000,
            "专业组最低分1": 595,
            "最低分1": 595,
            "最高分1": 610,
            "计划人数": 60,
            "专业组计划人数": 200,
            "院校排名": 120,
        },
    ]
    pilot_2025_rows = []
    for offset, row in enumerate(rows[:2], start=1):
        pilot_2025_rows.append(
            {
                **row,
                "ID": 100 + offset,
                "年份": 2025,
                "院校专业组代码": "10590225",
                "专业组名称": "物理225组",
                "专业组最低位次1": 7800,
                "最低位次1": 7600 + offset * 100,
                "专业组最低分1": 634,
                "最低分1": 632 - offset,
                "最高分1": 646 - offset,
            }
        )
    return pilot_2025_rows + rows


def _render_markdown(report: dict[str, Any]) -> str:
    lines = [
        "# Real Dataset Pilot 报告",
        "",
        f"- status：`{report.get('status')}`",
        f"- source_path：`{report['source_path']}`",
        f"- dataset_id：`{report['dataset_id']}`",
        f"- source_fingerprint：`{report.get('source_fingerprint')}`",
        f"- sheet_name：`{report.get('sheet_name')}`",
        f"- row_count / column_count：`{report.get('row_count')}` / `{report.get('column_count')}`",
        f"- detected_header_row：`{report.get('detected_header_row')}`",
        f"- warehouse_path：`{report.get('warehouse_path')}`",
        f"- warehouse_fingerprint：`{report.get('warehouse_fingerprint')}`",
        "",
        "## Schema Profile Summary",
        _json_block(report["schema_profile_summary"]),
        "## Risky Fields",
        _json_block(report["risky_fields"]),
        "## Approved / Blocked Fields",
        _json_block(
            {
                "approved_fields": report["approved_fields"],
                "blocked_fields": report["blocked_fields"],
                "required_fields": report["required_fields"],
                "missing_fields": report["missing_fields"],
            }
        ),
        "## Target Query Results",
        _json_block(report["target_query_results"]),
        "## Warnings",
        _json_block(report["warnings"]),
        "## Failures",
        _json_block(report["failures"]),
    ]
    return "\n".join(lines).rstrip() + "\n"


def _json_block(payload: Any) -> str:
    return "```json\n" + json.dumps(
        _json_ready(payload),
        ensure_ascii=False,
        indent=2,
    ) + "\n```"


def _dataset_id(source_path: Path) -> str:
    safe = "".join(char if char.isalnum() else "_" for char in source_path.stem.lower())
    safe = safe.strip("_")[:24] or "admissions"
    return f"ds_real_pilot_{safe}"


def _utc_now() -> str:
    return datetime.now(UTC).isoformat().replace("+00:00", "Z")


def _json_ready(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _json_ready(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_json_ready(item) for item in value]
    if isinstance(value, tuple):
        return [_json_ready(item) for item in value]
    if isinstance(value, Path):
        return str(value)
    if hasattr(value, "item"):
        try:
            return _json_ready(value.item())
        except (TypeError, ValueError):
            pass
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def _sanitize_report_paths(value: Any, *, base_dir: Path) -> Any:
    """净化报告中的路径字段，避免固化本机绝对路径。"""

    return _sanitize_report_path_value(value, base_dir=base_dir, key=None)


def _sanitize_report_path_value(
    value: Any,
    *,
    base_dir: Path,
    key: str | None,
) -> Any:
    if isinstance(value, dict):
        return {
            str(item_key): _sanitize_report_path_value(
                item,
                base_dir=base_dir,
                key=str(item_key),
            )
            for item_key, item in value.items()
        }
    if isinstance(value, list):
        return [
            _sanitize_report_path_value(item, base_dir=base_dir, key=key)
            for item in value
        ]
    if isinstance(value, Path):
        return _safe_report_path(value, base_dir=base_dir)
    if isinstance(value, str) and key in REPORT_PATH_KEYS:
        return _safe_report_path(value, base_dir=base_dir)
    return value


def _safe_report_path(value: str | Path, *, base_dir: Path) -> str:
    path = Path(value)
    if not path.is_absolute():
        return str(path)
    for root in (base_dir, ROOT_DIR):
        try:
            return str(path.resolve().relative_to(root.resolve()))
        except ValueError:
            continue
    return path.name


def _arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Run real admissions dataset upload/review/warehouse/query pilot."
    )
    parser.add_argument("source_path", nargs="?", help="CSV/Excel source path.")
    parser.add_argument("--fixture", action="store_true", help="Use built-in pilot fixture.")
    parser.add_argument("--sheet-name", default=None, help="Excel sheet name.")
    parser.add_argument("--output-dir", default=str(OUTPUT_DIR), help="Report output dir.")
    return parser


if __name__ == "__main__":
    raise SystemExit(main())
