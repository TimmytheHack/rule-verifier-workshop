"""从 CSV/Excel 自动生成 draft domain pack。"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from src.adapters.data_warehouse import build_structured_store_from_dataset
from src.adapters.excel_adapter import ExcelDataSet, cell_text
from src.domains import DomainConfig


DRAFT_STATUS = "draft"
REVIEW_STATUS = "needs_review"
SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xlsm", ".xls"}
PII_KEYWORDS = {
    "email",
    "e-mail",
    "mail",
    "phone",
    "mobile",
    "tel",
    "contact",
    "ssn",
    "passport",
    "id_card",
    "身份证",
    "电话",
    "手机",
    "邮箱",
    "邮件",
    "联系人",
}
IDENTIFIER_KEYWORDS = {
    "id",
    "code",
    "uuid",
    "编号",
    "代码",
    "编码",
}
NUMERIC_HINT_KEYWORDS = {
    "price",
    "rent",
    "cost",
    "fee",
    "tuition",
    "count",
    "score",
    "rank",
    "rating",
    "bedroom",
    "bath",
    "area",
    "year",
    "价格",
    "租金",
    "费用",
    "学费",
    "人数",
    "分数",
    "位次",
    "排名",
    "评分",
    "面积",
    "年份",
}
TEXT_HINT_KEYWORDS = {
    "name",
    "title",
    "description",
    "note",
    "remark",
    "address",
    "名称",
    "标题",
    "描述",
    "备注",
    "地址",
}
ASC_SORT_KEYWORDS = {
    "price",
    "rent",
    "cost",
    "fee",
    "tuition",
    "rank",
    "价格",
    "租金",
    "费用",
    "学费",
    "位次",
    "排名",
}
DESC_SORT_KEYWORDS = {
    "rating",
    "score",
    "评分",
    "分数",
}
ADMISSIONS_SCORE_KEYWORDS = {"分", "score"}
ADMISSIONS_RANK_KEYWORDS = {"位次", "rank"}
ADMISSIONS_GROUP_KEYWORDS = {"专业组", "group"}
ADMISSIONS_MAJOR_KEYWORDS = {"专业", "major"}
ADMISSIONS_UNIVERSITY_KEYWORDS = {"院校", "学校", "大学", "university", "school"}
ADMISSIONS_LOCATION_KEYWORDS = {"所在地", "所在省", "城市", "province", "city", "location"}
SPECIAL_PLAN_TERMS = [
    "中外合作",
    "国际班",
    "境外培养",
    "合作办学",
    "校企合作",
    "地方专项",
    "专项计划",
]
HEADER_SCAN_ROWS = 30
FORMULA_SCAN_LIMIT = 5000


@dataclass(frozen=True)
class DomainPackGenerationResult:
    """生成结果摘要。"""

    domain_dir: Path
    domain_config_path: Path
    schema_path: Path
    database_path: Path
    schema_profile_path: Path
    value_index_path: Path
    ingestion_summary_path: Path
    row_count: int
    column_count: int

    def to_dict(self) -> dict[str, Any]:
        return {
            "domain_dir": str(self.domain_dir),
            "domain_config_path": str(self.domain_config_path),
            "schema_path": str(self.schema_path),
            "database_path": str(self.database_path),
            "schema_profile_path": str(self.schema_profile_path),
            "value_index_path": str(self.value_index_path),
            "ingestion_summary_path": str(self.ingestion_summary_path),
            "row_count": self.row_count,
            "column_count": self.column_count,
        }


@dataclass(frozen=True)
class SourceInspection:
    """CSV/Excel 读取审查结果。"""

    dataset: ExcelDataSet
    sheet_summaries: list[dict[str, Any]]
    warnings: list[dict[str, Any]]
    detected_header_row: int
    header_detection_status: str
    original_column_mapping: list[dict[str, Any]]

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet_name": self.dataset.sheet_name,
            "sheet_summaries": self.sheet_summaries,
            "warnings": self.warnings,
            "detected_header_row": self.detected_header_row,
            "header_detection_status": self.header_detection_status,
            "original_column_mapping": self.original_column_mapping,
            "row_count": len(self.dataset.dataframe),
            "column_count": len(self.dataset.headers),
        }


def generate_domain_pack(
    source_path: str | Path,
    domain_name: str,
    output_root: str | Path = ROOT_DIR / "domains",
    llm: str = "off",
    sheet_name: str | None = None,
) -> DomainPackGenerationResult:
    """生成 draft domain pack，并复用 DuckDB warehouse ingestion。"""

    if llm not in {"off", "deepseek"}:
        raise ValueError("--llm must be one of: off, deepseek")
    domain_id = _normalize_domain_name(domain_name)
    dataset = load_source_dataset(source_path, sheet_name=sheet_name)
    profile = profile_dataset(dataset, domain_id)
    domain_dir = Path(output_root) / domain_id
    domain_dir.mkdir(parents=True, exist_ok=True)

    llm_candidates = _llm_candidates(profile, llm) if llm != "off" else {}
    payloads = build_draft_payloads(
        domain_id=domain_id,
        source_path=Path(source_path),
        profile=profile,
        llm_candidates=llm_candidates,
    )
    _write_payloads(domain_dir, payloads)

    database_path = domain_dir / f"{domain_id}.duckdb"
    value_index_path = domain_dir / "schema_value_index.json"
    ingestion = build_structured_store_from_dataset(
        dataset=dataset,
        schema_path=domain_dir / "schema_registry.json",
        database_path=database_path,
        index_path=value_index_path,
        table_name=domain_id,
        source_path=source_path,
    )
    ingestion_summary_path = domain_dir / "ingestion_summary.json"
    _write_json(ingestion_summary_path, ingestion.to_dict())

    return DomainPackGenerationResult(
        domain_dir=domain_dir,
        domain_config_path=domain_dir / "domain.json",
        schema_path=domain_dir / "schema_registry.json",
        database_path=database_path,
        schema_profile_path=domain_dir / "schema_profile.json",
        value_index_path=value_index_path,
        ingestion_summary_path=ingestion_summary_path,
        row_count=ingestion.row_count,
        column_count=ingestion.column_count,
    )


def load_source_dataset(
    source_path: str | Path,
    sheet_name: str | None = None,
) -> ExcelDataSet:
    """读取 CSV/Excel 为统一 dataset 结构。"""

    return inspect_source_dataset(source_path, sheet_name=sheet_name).dataset


def inspect_source_dataset(
    source_path: str | Path,
    sheet_name: str | None = None,
) -> SourceInspection:
    """读取并审查 CSV/Excel，包含 sheet、header 和字段清洗信息。"""

    path = Path(source_path)
    if not path.exists():
        raise FileNotFoundError(f"Source file not found: {path}")
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported source extension: {suffix}")
    if suffix == ".csv":
        return _inspect_csv_source(path)
    return _inspect_excel_source(path, sheet_name=sheet_name)


def _inspect_csv_source(path: Path) -> SourceInspection:
    raw = pd.read_csv(path, header=None, dtype=object)
    sheet_summaries = [
        _sheet_summary(
            sheet_name=path.stem,
            row_count=len(raw),
            column_count=len(raw.columns),
            non_empty_cells=_non_empty_cell_count(raw),
            selected=True,
        )
    ]
    header = _detect_header_row_from_frame(raw)
    dataframe = _dataframe_from_raw(raw, header, source_name=path.stem)
    return _inspection_from_dataframe(
        path=path,
        sheet_name=path.stem,
        dataframe=dataframe,
        header=header,
        sheet_summaries=sheet_summaries,
        warnings=[],
    )


def _inspect_excel_source(
    path: Path,
    sheet_name: str | None,
) -> SourceInspection:
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        sheet_summaries = [
            _openpyxl_sheet_summary(sheet, selected=False)
            for sheet in workbook.worksheets
        ]
        selected_sheet = sheet_name or _first_non_empty_sheet(sheet_summaries)
        if selected_sheet not in workbook.sheetnames:
            raise ValueError(
                f"Sheet not found: {selected_sheet}. "
                f"Available sheets: {', '.join(workbook.sheetnames)}"
            )
        selected_summary = []
        for summary in sheet_summaries:
            updated = dict(summary)
            updated["selected"] = summary["sheet_name"] == selected_sheet
            selected_summary.append(updated)
        sheet = workbook[selected_sheet]
        warnings = _excel_structural_warnings(sheet)
        raw = pd.read_excel(
            path,
            sheet_name=selected_sheet,
            header=None,
            dtype=object,
            engine="openpyxl",
        )
    finally:
        workbook.close()
    header = _detect_header_row_from_frame(raw)
    dataframe = _dataframe_from_raw(raw, header, source_name=selected_sheet)
    return _inspection_from_dataframe(
        path=path,
        sheet_name=selected_sheet,
        dataframe=dataframe,
        header=header,
        sheet_summaries=selected_summary,
        warnings=warnings,
    )


def _inspection_from_dataframe(
    *,
    path: Path,
    sheet_name: str,
    dataframe: pd.DataFrame,
    header: dict[str, Any],
    sheet_summaries: list[dict[str, Any]],
    warnings: list[dict[str, Any]],
) -> SourceInspection:
    normalized = _normalize_dataframe_columns(dataframe)
    cleaned = normalized["dataframe"]
    header_warnings = list(header.get("warnings") or [])
    warnings = [*warnings, *header_warnings, *normalized["warnings"]]
    dataset = ExcelDataSet(
        workbook_path=path,
        sheet_name=sheet_name,
        header_row=header["row_number"],
        headers=list(cleaned.columns),
        header_index={name: index for index, name in enumerate(cleaned.columns)},
        dataframe=cleaned,
        sheet_summaries=sheet_summaries,
        warnings=warnings,
        header_detection_status=header["status"],
        original_column_mapping=normalized["mapping"],
    )
    return SourceInspection(
        dataset=dataset,
        sheet_summaries=sheet_summaries,
        warnings=warnings,
        detected_header_row=header["row_number"],
        header_detection_status=header["status"],
        original_column_mapping=normalized["mapping"],
    )


def _dataframe_from_raw(
    raw: pd.DataFrame,
    header: dict[str, Any],
    *,
    source_name: str,
) -> pd.DataFrame:
    _ = source_name
    header_index = max(0, int(header["row_number"]) - 1)
    headers = [
        _normalize_column_name(value)
        for value in raw.iloc[header_index].tolist()
    ]
    data = raw.iloc[header_index + 1 :].copy()
    data.columns = headers
    data = data.dropna(how="all")
    return data


def _detect_header_row_from_frame(raw: pd.DataFrame) -> dict[str, Any]:
    candidates = []
    max_rows = min(len(raw), HEADER_SCAN_ROWS)
    for index in range(max_rows):
        values = raw.iloc[index].tolist()
        score = _header_score(values)
        candidates.append(
            {
                "row_number": index + 1,
                "score": score,
                "non_empty": sum(1 for value in values if cell_text(value)),
            }
        )
    candidates.sort(key=lambda item: item["score"], reverse=True)
    if not candidates:
        return {
            "row_number": 1,
            "status": "needs_review",
            "confidence": 0,
            "warnings": [_source_warning("header_row_not_found", "未找到可用表头行。")],
        }
    best = candidates[0]
    second_score = candidates[1]["score"] if len(candidates) > 1 else -1
    status = "confirmed"
    warnings = []
    if best["score"] < 3 or best["score"] - second_score < 1:
        status = "needs_review"
        warnings.append(
            _source_warning(
                "header_row_detection_needs_review",
                "表头行检测置信度不足，请人工确认。",
                detected_header_row=best["row_number"],
                confidence=best["score"],
            )
        )
    return {
        "row_number": int(best["row_number"]),
        "status": status,
        "confidence": best["score"],
        "warnings": warnings,
    }


def _header_score(values: list[Any]) -> float:
    texts = [_normalize_column_name(value) for value in values]
    non_empty = [text for text in texts if text]
    if not non_empty:
        return 0
    unique_count = len(set(non_empty))
    numeric_count = sum(1 for text in non_empty if _parse_number(text) is not None)
    keyword_count = sum(1 for text in non_empty if _header_keyword_hit(text))
    avg_length = sum(len(text) for text in non_empty) / len(non_empty)
    duplicate_penalty = max(0, len(non_empty) - unique_count)
    length_penalty = 1 if avg_length > 40 else 0
    return (
        len(non_empty)
        + keyword_count * 2
        - numeric_count * 0.75
        - duplicate_penalty
        - length_penalty
    )


def _header_keyword_hit(text: str) -> bool:
    lowered = text.lower()
    keywords = (
        NUMERIC_HINT_KEYWORDS
        | TEXT_HINT_KEYWORDS
        | IDENTIFIER_KEYWORDS
        | ADMISSIONS_SCORE_KEYWORDS
        | ADMISSIONS_RANK_KEYWORDS
        | ADMISSIONS_UNIVERSITY_KEYWORDS
        | ADMISSIONS_LOCATION_KEYWORDS
    )
    return any(keyword.lower() in lowered for keyword in keywords)


def _normalize_dataframe_columns(dataframe: pd.DataFrame) -> dict[str, Any]:
    original_columns = list(dataframe.columns)
    normalized = [_normalize_column_name(column) for column in original_columns]
    safe_columns, duplicate_warnings = _deduplicate_columns(normalized)
    mapping = [
        {
            "column_index": index,
            "original_column": cell_text(original),
            "normalized_column": normalized_name,
            "safe_column": safe_name,
        }
        for index, (original, normalized_name, safe_name) in enumerate(
            zip(original_columns, normalized, safe_columns, strict=True),
            start=1,
        )
    ]
    cleaned = dataframe.copy()
    cleaned.columns = safe_columns
    drop_columns = [
        column
        for column, original in zip(cleaned.columns, normalized, strict=True)
        if not original or _column_is_empty(cleaned[column])
    ]
    warnings = duplicate_warnings
    if drop_columns:
        warnings.append(
            _source_warning(
                "empty_columns_dropped",
                "已丢弃空列或无表头列。",
                columns=drop_columns,
            )
        )
        cleaned = cleaned.drop(columns=drop_columns)
        mapping = [item for item in mapping if item["safe_column"] not in drop_columns]
    cleaned = cleaned.dropna(how="all")
    cleaned = _coerce_numeric_columns(cleaned)
    cleaned = cleaned.reset_index(drop=True)
    return {"dataframe": cleaned, "mapping": mapping, "warnings": warnings}


def _coerce_numeric_columns(dataframe: pd.DataFrame) -> pd.DataFrame:
    """把纯数值列恢复为 numeric dtype，避免 CSV object 读取污染 profile。"""

    coerced = dataframe.copy()
    for column in coerced.columns:
        series = coerced[column]
        non_empty = [
            value
            for value in series.tolist()
            if cell_text(value)
        ]
        if not non_empty:
            continue
        numeric = pd.to_numeric(series, errors="coerce")
        non_empty_numeric = pd.to_numeric(pd.Series(non_empty), errors="coerce")
        if not non_empty_numeric.isna().any():
            coerced[column] = numeric
    return coerced


def _normalize_column_name(value: Any) -> str:
    text = cell_text(value)
    text = text.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _deduplicate_columns(columns: list[str]) -> tuple[list[str], list[dict[str, Any]]]:
    counts: Counter[str] = Counter()
    duplicates = []
    safe_columns = []
    for column in columns:
        if not column:
            safe_columns.append("")
            continue
        counts[column] += 1
        safe = column if counts[column] == 1 else f"{column}_{counts[column]}"
        if counts[column] == 2:
            duplicates.append(column)
        safe_columns.append(safe)
    warnings = []
    if duplicates:
        warnings.append(
            _source_warning(
                "duplicate_columns_normalized",
                "检测到重复列名，已自动生成安全列名并保留原始映射。",
                columns=duplicates,
            )
        )
    return safe_columns, warnings


def _column_is_empty(series: pd.Series) -> bool:
    return all(not cell_text(value) for value in series.dropna().tolist())


def _openpyxl_sheet_summary(sheet: Any, selected: bool) -> dict[str, Any]:
    non_empty = 0
    for row in sheet.iter_rows(values_only=True):
        for value in row:
            if cell_text(value):
                non_empty += 1
    return _sheet_summary(
        sheet_name=sheet.title,
        row_count=sheet.max_row or 0,
        column_count=sheet.max_column or 0,
        non_empty_cells=non_empty,
        selected=selected,
    )


def _sheet_summary(
    *,
    sheet_name: str,
    row_count: int,
    column_count: int,
    non_empty_cells: int,
    selected: bool,
) -> dict[str, Any]:
    return {
        "sheet_name": sheet_name,
        "row_count": int(row_count),
        "column_count": int(column_count),
        "non_empty_cells": int(non_empty_cells),
        "selected": selected,
    }


def _first_non_empty_sheet(sheet_summaries: list[dict[str, Any]]) -> str:
    for summary in sheet_summaries:
        if int(summary.get("non_empty_cells") or 0) > 0:
            return str(summary["sheet_name"])
    if sheet_summaries:
        return str(sheet_summaries[0]["sheet_name"])
    raise ValueError("Workbook has no sheets.")


def _excel_structural_warnings(sheet: Any) -> list[dict[str, Any]]:
    warnings = []
    if sheet.merged_cells.ranges:
        merged_ranges = list(sheet.merged_cells.ranges)
        warnings.append(
            _source_warning(
                "merged_cells_detected",
                "检测到合并单元格，请确认表头和数据未被合并结构影响。",
                ranges=[str(item) for item in merged_ranges[:20]],
            )
        )
    hidden_rows = [
        index for index, dimension in sheet.row_dimensions.items() if dimension.hidden
    ]
    hidden_columns = [
        key for key, dimension in sheet.column_dimensions.items() if dimension.hidden
    ]
    if hidden_rows or hidden_columns:
        warnings.append(
            _source_warning(
                "hidden_rows_or_columns_detected",
                "检测到隐藏行或隐藏列，ingestion 会按文件内容读取。",
                hidden_rows=hidden_rows[:20],
                hidden_columns=hidden_columns[:20],
            )
        )
    formula_cells = []
    scanned = 0
    for row in sheet.iter_rows():
        for cell in row:
            scanned += 1
            if cell.data_type == "f":
                formula_cells.append(cell.coordinate)
            if scanned >= FORMULA_SCAN_LIMIT:
                break
        if scanned >= FORMULA_SCAN_LIMIT:
            break
    if formula_cells:
        warnings.append(
            _source_warning(
                "formula_cells_detected",
                "检测到公式单元格；读取值可能依赖 Excel 缓存，请人工核验。",
                cells=formula_cells[:20],
            )
        )
    return warnings


def _non_empty_cell_count(dataframe: pd.DataFrame) -> int:
    count = 0
    for value in dataframe.to_numpy().ravel().tolist():
        if cell_text(value):
            count += 1
    return count


def _source_warning(
    code: str,
    message: str,
    **details: Any,
) -> dict[str, Any]:
    return {"code": code, "severity": "warning", "message": message, **details}


def profile_dataset(dataset: ExcelDataSet, domain_id: str) -> dict[str, Any]:
    """基于列事实生成通用 schema profile。"""

    dataframe = dataset.dataframe
    row_count = len(dataframe)
    field_ids = _field_ids(dataset.headers)
    columns = []
    for index, source_column in enumerate(dataset.headers, start=1):
        if source_column not in dataframe.columns:
            continue
        series = dataframe[source_column]
        non_empty = _non_empty_values(series)
        unique_count = int(pd.Series(non_empty).nunique(dropna=True)) if non_empty else 0
        numeric_values = [_parse_number(value) for value in non_empty[:1000]]
        numeric_count = sum(value is not None for value in numeric_values)
        numeric_ratio = numeric_count / min(len(non_empty), 1000) if non_empty else 0.0
        keyword_flags = _keyword_flags(source_column)
        pii = _has_keyword(source_column, PII_KEYWORDS)
        raw_high_cardinality = _is_high_cardinality(unique_count, len(non_empty))
        inferred_type = _infer_type(
            series=series,
            non_empty=non_empty,
            unique_count=unique_count,
            numeric_ratio=numeric_ratio,
            keyword_flags=keyword_flags,
        )
        high_cardinality = raw_high_cardinality and inferred_type not in {
            "number",
            "number_from_string",
        }
        role = _infer_role(
            source_column=source_column,
            inferred_type=inferred_type,
            high_cardinality=high_cardinality,
            pii=pii,
        )
        numeric_profile = _numeric_profile(non_empty)
        samples = _sample_values(non_empty, pii=pii)
        admissions_semantics = _admissions_semantics(source_column)
        special_plan_risk = _special_plan_risk(source_column, samples)
        columns.append(
            {
                "column_index": index,
                "source_column": source_column,
                "field_id": field_ids[source_column],
                "label": source_column,
                "dtype": str(series.dtype),
                "inferred_type": inferred_type,
                "role": role,
                "status": REVIEW_STATUS,
                "null_rate": round(
                    1 - (len(non_empty) / row_count), 4
                )
                if row_count
                else 0.0,
                "non_empty_count": len(non_empty),
                "unique_count": unique_count,
                "unique_rate": round(unique_count / len(non_empty), 4)
                if non_empty
                else 0.0,
                "high_cardinality": high_cardinality,
                "pii_risk": pii,
                "sample_values": samples,
                "numeric": numeric_profile,
                "keyword_flags": keyword_flags,
                "admissions_semantics": admissions_semantics,
                "special_plan_risk": special_plan_risk,
                "candidate_allowed_ops": _candidate_allowed_ops(
                    inferred_type=inferred_type,
                    role=role,
                    high_cardinality=high_cardinality,
                    pii=pii,
                ),
                "filter_policy": "blocked_by_default"
                if pii or high_cardinality
                else REVIEW_STATUS,
            }
        )
    return {
        "domain_id": domain_id,
        "status": DRAFT_STATUS,
        "source": {
            "path": str(dataset.workbook_path),
            "sheet_name": dataset.sheet_name,
            "header_row": dataset.header_row,
            "detected_header_row": dataset.header_row,
            "header_detection_status": dataset.header_detection_status,
            "sheet_summaries": dataset.sheet_summaries or [],
            "warnings": dataset.warnings or [],
            "original_column_mapping": dataset.original_column_mapping or [],
        },
        "row_count": row_count,
        "column_count": len(columns),
        "columns": columns,
        "methodology": {
            "basis": [
                "列名",
                "dtype",
                "空值率",
                "唯一值数量",
                "样例值",
                "数值范围",
                "字段名关键词",
            ],
            "safety": (
                "自动生成结果只作为 draft/needs_review；字段必须人工 approve "
                "后才会获得 allowed_ops。"
            ),
        },
}


def _admissions_semantics(source_column: str) -> dict[str, Any]:
    text = source_column.lower()
    score = _has_any(text, ADMISSIONS_SCORE_KEYWORDS)
    rank = _has_any(text, ADMISSIONS_RANK_KEYWORDS)
    group = _has_any(text, ADMISSIONS_GROUP_KEYWORDS)
    major = _has_any(text, ADMISSIONS_MAJOR_KEYWORDS)
    minimum = any(term in text for term in ["最低", "min"])
    semantics = {
        "score_kind": None,
        "rank_kind": None,
        "entity_kind": None,
        "location_kind": None,
    }
    if score:
        if group and minimum:
            semantics["score_kind"] = "group_min_score"
        elif major and minimum:
            semantics["score_kind"] = "major_min_score"
        elif minimum:
            semantics["score_kind"] = "min_score"
        else:
            semantics["score_kind"] = "score"
    if rank:
        if group and minimum:
            semantics["rank_kind"] = "group_min_rank"
        elif major and minimum:
            semantics["rank_kind"] = "major_min_rank"
        elif minimum:
            semantics["rank_kind"] = "min_rank"
        else:
            semantics["rank_kind"] = "rank"
    if any(term in text for term in ["院校名称", "学校名称", "university_name"]):
        semantics["entity_kind"] = "university_name"
    elif any(term in text for term in ["学院", "college"]):
        semantics["entity_kind"] = "college_name"
    if any(term in text for term in ["所在省", "学校所在地", "school_location"]):
        semantics["location_kind"] = "school_location"
    elif any(term in text for term in ["生源地", "录取省份", "admission_province"]):
        semantics["location_kind"] = "admission_province"
    return {key: value for key, value in semantics.items() if value}


def _special_plan_risk(source_column: str, samples: list[Any]) -> dict[str, Any]:
    haystack = " ".join([source_column, *[str(value) for value in samples]])
    matched = [term for term in SPECIAL_PLAN_TERMS if term in haystack]
    return {
        "matched_terms": matched,
        "needs_schema_approval": bool(matched),
    }


def _has_any(text: str, keywords: set[str]) -> bool:
    return any(keyword.lower() in text for keyword in keywords)


def build_draft_payloads(
    domain_id: str,
    source_path: Path,
    profile: dict[str, Any],
    llm_candidates: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """构造 draft 配置文件内容。"""

    llm_candidates = llm_candidates or {}
    fields = {
        column["field_id"]: _draft_schema_field(column)
        for column in profile["columns"]
    }
    top_result_candidates = _top_result_candidates(profile["columns"])
    sort_candidates = _sort_candidates(profile["columns"])
    extraction_aliases = _extraction_aliases(profile["columns"], llm_candidates)

    domain_json = {
        "domain_id": domain_id,
        "display_name": domain_id,
        "status": DRAFT_STATUS,
        "review_required": True,
        "data": {
            "workbook_path": str(source_path),
            "fixture_path": str(source_path),
            "table_name": domain_id,
            "warehouse_database_path": f"{domain_id}.duckdb",
            "value_index_path": "schema_value_index.json",
            "required_field_ids": [],
        },
        "paths": {
            "schema": "schema_registry.json",
            "attribute_grounding": "attribute_grounding.json",
            "rule_taxonomy": "rule_taxonomy.json",
            "value_aliases": "value_aliases.json",
            "answer_templates": "answer_templates.json",
            "golden_cases": "golden_cases.json",
        },
        "execution": _empty_execution_config(),
        "top_result_mapping": [],
        "workbench": {
            "slot_labels": {},
            "extracted_preferences": [],
            "reviewed_candidate_mappings": [],
            "not_executed_overrides": [],
            "context_warnings": [
                "该 domain pack 由自动生成器创建，字段和规则都需要人工 review 后才能执行。"
            ],
        },
    }
    rule_taxonomy_json = {
        "status": DRAFT_STATUS,
        "deterministic_rules": [],
        "context_rules": [],
        "candidate_rules": [],
        "llm_needed_parts": [],
        "confirmation_questions": [],
        "simulated_confirmations": {},
        "confirmed_candidate_rules": [],
        "non_executable_preferences": [],
        "draft_rule_candidates": _rule_candidates(profile["columns"]),
    }
    answer_templates_json = {
        "status": DRAFT_STATUS,
        "rank_field_id": None,
        "money_field_ids": [],
        "result_line_fields": [],
        "result_text_fields": [],
    }
    golden_cases_json = {
        "status": DRAFT_STATUS,
        "cases": [],
        "notes": "自动生成的 golden cases seed 为空，需要人工补充并 review。",
    }
    value_aliases_json = {
        "status": DRAFT_STATUS,
        "description": "自动生成候选 aliases，只能作为 review 输入。",
        "candidate_aliases": extraction_aliases["fields"],
    }
    attribute_grounding_json = {
        "status": DRAFT_STATUS,
        "description": "自动生成占位 grounding policy，默认不把任何 slot 接地为可执行属性。",
        "slot_policies": {},
        "other_vague_policies": {},
    }

    domain_yaml = {
        "status": DRAFT_STATUS,
        "domain_id": domain_id,
        "runtime_config": "domain.json",
        "source": profile["source"],
        "warehouse": {
            "database_path": f"{domain_id}.duckdb",
            "value_index_path": "schema_value_index.json",
            "table_name": domain_id,
        },
        "review_required": True,
        "notes": [
            "该文件由 schema profiling 自动生成。",
            "所有字段和规则默认不可执行，必须人工 approve。",
        ],
    }
    schema_mapping_yaml = {
        "status": DRAFT_STATUS,
        "domain_id": domain_id,
        "fields": [
            {
                "field_id": column["field_id"],
                "source_column": column["source_column"],
                "label": column["label"],
                "inferred_type": column["inferred_type"],
                "role": column["role"],
                "status": REVIEW_STATUS,
                "filter_policy": column["filter_policy"],
                "candidate_allowed_ops": column["candidate_allowed_ops"],
            }
            for column in profile["columns"]
        ],
    }
    rule_taxonomy_seed = {
        "status": DRAFT_STATUS,
        "domain_id": domain_id,
        "deterministic_rules": [],
        "candidate_rules": _rule_candidates(profile["columns"]),
        "notes": [
            "seed 只提出候选规则形状。",
            "不能直接进入 RuleVerifier hard rules。",
        ],
    }
    sort_policy_seed = {
        "status": DRAFT_STATUS,
        "domain_id": domain_id,
        "candidate_sort_policy": sort_candidates,
    }
    top_result_mapping_yaml = {
        "status": DRAFT_STATUS,
        "domain_id": domain_id,
        "candidate_top_result_mapping": top_result_candidates,
    }
    answer_templates_seed = _answer_template_seed(
        domain_id=domain_id,
        top_result_candidates=top_result_candidates,
        llm_candidates=llm_candidates,
    )
    golden_cases_seed = {
        "status": DRAFT_STATUS,
        "domain_id": domain_id,
        "cases": [],
        "notes": [
            "自动生成器不能替代人工标注 golden cases。",
            "新增 case 后仍需运行 pipeline 验证。",
        ],
    }

    return {
        "json": {
            "domain.json": domain_json,
            "schema_registry.json": {"status": DRAFT_STATUS, "fields": fields},
            "rule_taxonomy.json": rule_taxonomy_json,
            "answer_templates.json": answer_templates_json,
            "golden_cases.json": golden_cases_json,
            "value_aliases.json": value_aliases_json,
            "attribute_grounding.json": attribute_grounding_json,
            "schema_profile.json": profile,
            "extraction_aliases.seed.json": extraction_aliases,
        },
        "yaml": {
            "domain.yaml": domain_yaml,
            "schema_mapping.yaml": schema_mapping_yaml,
            "rule_taxonomy.seed.yaml": rule_taxonomy_seed,
            "top_result_mapping.yaml": top_result_mapping_yaml,
            "sort_policy.seed.yaml": sort_policy_seed,
            "golden_cases.seed.yaml": golden_cases_seed,
        },
        "text": {
            "answer_templates.seed.md": answer_templates_seed,
        },
    }


def approve_domain_pack(
    domain_dir: str | Path,
    approved_field_ids: list[str] | None = None,
    output_field_ids: list[str] | None = None,
    sort_field_id: str | None = None,
) -> DomainConfig:
    """人工 review 后启用指定字段，测试和小型 toy domain 可直接复用。"""

    root = Path(domain_dir)
    schema_path = root / "schema_registry.json"
    domain_path = root / "domain.json"
    schema = json.loads(schema_path.read_text(encoding="utf-8"))
    domain = json.loads(domain_path.read_text(encoding="utf-8"))
    fields = schema["fields"]
    approved = approved_field_ids or list(fields)
    output_fields = output_field_ids or approved
    approved_set = set(approved)

    for field_id, spec in fields.items():
        if field_id not in approved_set:
            spec["status"] = REVIEW_STATUS
            spec["allowed_ops"] = []
            continue
        spec["status"] = "active"
        spec["reviewed"] = True
        spec["allowed_ops"] = list(spec.get("candidate_allowed_ops") or [])

    domain["status"] = "approved"
    domain["review_required"] = False
    domain["data"]["required_field_ids"] = [
        field_id for field_id in output_fields if field_id in fields
    ]
    domain["execution"] = _approved_execution_config(
        fields=fields,
        output_field_ids=output_fields,
        sort_field_id=sort_field_id,
    )
    domain["top_result_mapping"] = [
        {"key": field_id, "field_id": field_id}
        for field_id in output_fields
        if field_id in fields
    ]
    _write_json(schema_path, schema)
    _write_json(domain_path, domain)
    _write_json(
        root / "answer_templates.json",
        _approved_answer_templates(fields, output_fields),
    )
    return DomainConfig.from_path(root)


def _write_payloads(domain_dir: Path, payloads: dict[str, Any]) -> None:
    for filename, payload in payloads["json"].items():
        _write_json(domain_dir / filename, payload)
    for filename, payload in payloads["yaml"].items():
        (domain_dir / filename).write_text(_to_yaml(payload), encoding="utf-8")
    for filename, content in payloads["text"].items():
        (domain_dir / filename).write_text(content, encoding="utf-8")


def _draft_schema_field(column: dict[str, Any]) -> dict[str, Any]:
    return {
        "source_column": column["source_column"],
        "type": _registry_type(column["inferred_type"]),
        "label": column["label"],
        "aliases": _field_aliases(column),
        "allowed_ops": [],
        "candidate_allowed_ops": column["candidate_allowed_ops"],
        "nullable": column["null_rate"] > 0,
        "status": REVIEW_STATUS,
        "review_required": True,
        "filter_policy": column["filter_policy"],
        "role": column["role"],
        "profile": {
            "dtype": column["dtype"],
            "null_rate": column["null_rate"],
            "unique_count": column["unique_count"],
            "sample_values": column["sample_values"],
            "numeric": column["numeric"],
            "keyword_flags": column["keyword_flags"],
            "high_cardinality": column["high_cardinality"],
            "pii_risk": column["pii_risk"],
        },
        "notes": "自动生成 draft 字段，人工 review 前不能执行。",
    }


def _empty_execution_config() -> dict[str, Any]:
    return {
        "rank_field_id": None,
        "tuition_field_id": None,
        "projectable_required_field_ids": [],
        "numeric_helper_fields": [],
        "sort_policy": [],
        "row_number_output_key": "source_row_number",
        "row_number_offset": 1,
        "output_fields": [],
        "static_output_fields": {},
    }


def _approved_execution_config(
    fields: dict[str, dict[str, Any]],
    output_field_ids: list[str],
    sort_field_id: str | None,
) -> dict[str, Any]:
    numeric_helpers = []
    output_fields = []
    helper_by_field: dict[str, str] = {}
    for field_id in output_field_ids:
        spec = fields.get(field_id)
        if not spec:
            continue
        transform = _transform_for_type(spec.get("type"))
        output_spec = {
            "output_key": field_id,
            "field_id": field_id,
            "transform": transform,
        }
        if spec.get("type") in {"number", "number_from_string"}:
            helper_name = f"__{field_id}_num"
            helper_by_field[field_id] = helper_name
            numeric_helpers.append({"name": helper_name, "field_id": field_id})
            output_spec["helper"] = helper_name
        output_fields.append(output_spec)

    sort_policy = []
    chosen_sort = sort_field_id or _first_sortable_field(fields, output_field_ids)
    if chosen_sort and chosen_sort in helper_by_field:
        sort_policy.append(
            {
                "helper": helper_by_field[chosen_sort],
                "label_field_id": chosen_sort,
                "direction": _sort_direction(fields[chosen_sort]["source_column"]),
                "nulls": "LAST",
            }
        )
    return {
        "rank_field_id": None,
        "tuition_field_id": None,
        "projectable_required_field_ids": [],
        "numeric_helper_fields": numeric_helpers,
        "sort_policy": sort_policy,
        "row_number_output_key": "source_row_number",
        "row_number_offset": 1,
        "output_fields": output_fields,
        "static_output_fields": {},
    }


def _approved_answer_templates(
    fields: dict[str, dict[str, Any]],
    output_field_ids: list[str],
) -> dict[str, Any]:
    result_fields = []
    for field_id in output_field_ids:
        spec = fields.get(field_id)
        if not spec:
            continue
        result_fields.append(
            {
                "field_id": field_id,
                "label": spec.get("label") or field_id,
                "evidence_key": field_id,
                "format": "money" if _money_like(spec.get("source_column")) else None,
            }
        )
    return {
        "rank_field_id": None,
        "money_field_ids": [
            field_id
            for field_id in output_field_ids
            if _money_like((fields.get(field_id) or {}).get("source_column"))
        ],
        "result_line_fields": result_fields,
        "result_text_fields": result_fields,
    }


def _field_ids(headers: list[str]) -> dict[str, str]:
    used: set[str] = set()
    result = {}
    for index, header in enumerate(headers, start=1):
        base = _slugify(header) or f"field_{index:02d}"
        candidate = base
        suffix = 2
        while candidate in used:
            candidate = f"{base}_{suffix}"
            suffix += 1
        used.add(candidate)
        result[header] = candidate
    return result


def _slugify(value: str) -> str:
    text = re.sub(r"[^0-9A-Za-z]+", "_", value.strip().lower())
    text = re.sub(r"_+", "_", text).strip("_")
    if text and text[0].isdigit():
        text = f"field_{text}"
    return text


def _non_empty_values(series: pd.Series) -> list[Any]:
    values = []
    for value in series.dropna().tolist():
        text = cell_text(value)
        if text:
            values.append(value)
    return values


def _sample_values(values: list[Any], pii: bool, max_samples: int = 6) -> list[Any]:
    samples = []
    seen = set()
    for value in values:
        text = cell_text(value)
        if not text or text in seen:
            continue
        seen.add(text)
        samples.append(_redact_sample(text) if pii else _json_scalar(value))
        if len(samples) >= max_samples:
            break
    return samples


def _infer_type(
    series: pd.Series,
    non_empty: list[Any],
    unique_count: int,
    numeric_ratio: float,
    keyword_flags: dict[str, bool],
) -> str:
    if not non_empty:
        return "mostly_empty"
    if pd.api.types.is_numeric_dtype(series) or numeric_ratio >= 0.95:
        return "number" if pd.api.types.is_numeric_dtype(series) else "number_from_string"
    max_length = max((len(cell_text(value)) for value in non_empty), default=0)
    avg_length = sum(len(cell_text(value)) for value in non_empty) / len(non_empty)
    category_limit = min(30, max(5, int(len(non_empty) * 0.4)))
    if unique_count <= category_limit and max_length <= 60:
        return "enum"
    if max_length > 120 or avg_length > 50 or keyword_flags["text_hint"]:
        return "long_text"
    return "string"


def _infer_role(
    source_column: str,
    inferred_type: str,
    high_cardinality: bool,
    pii: bool,
) -> str:
    if pii:
        return "pii"
    if _has_keyword(source_column, IDENTIFIER_KEYWORDS):
        return "identifier"
    if inferred_type in {"number", "number_from_string"}:
        return "metric"
    if inferred_type == "enum":
        return "category"
    if high_cardinality:
        return "high_cardinality_text"
    return "text"


def _keyword_flags(source_column: str) -> dict[str, bool]:
    return {
        "numeric_hint": _has_keyword(source_column, NUMERIC_HINT_KEYWORDS),
        "identifier_hint": _has_keyword(source_column, IDENTIFIER_KEYWORDS),
        "text_hint": _has_keyword(source_column, TEXT_HINT_KEYWORDS),
        "pii_hint": _has_keyword(source_column, PII_KEYWORDS),
    }


def _has_keyword(source_column: Any, keywords: set[str]) -> bool:
    text = str(source_column).lower()
    tokens = set(re.split(r"[^0-9a-zA-Z一-龥]+", text))
    for keyword in keywords:
        lowered = keyword.lower()
        if lowered in tokens or lowered in text:
            return True
    return False


def _is_high_cardinality(unique_count: int, non_empty_count: int) -> bool:
    if non_empty_count < 10:
        return False
    return unique_count >= 10 and unique_count / non_empty_count >= 0.8


def _candidate_allowed_ops(
    inferred_type: str,
    role: str,
    high_cardinality: bool,
    pii: bool,
) -> list[str]:
    if pii or high_cardinality:
        return []
    if role == "identifier":
        return ["eq", "in", "sort"]
    if inferred_type in {"number", "number_from_string"}:
        return ["eq", "<=", ">=", "between", "sort"]
    if inferred_type == "enum":
        return ["eq", "in", "not_in"]
    if role == "text":
        return ["contains", "contains_any"]
    return []


def _numeric_profile(values: list[Any]) -> dict[str, Any] | None:
    parsed = [_parse_number(value) for value in values]
    numbers = [value for value in parsed if value is not None]
    if not numbers:
        return None
    return {"min": _clean_number(min(numbers)), "max": _clean_number(max(numbers))}


def _parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"-?\d+(?:\.\d+)?", str(value).replace(",", ""))
    return float(match.group()) if match else None


def _clean_number(value: float) -> int | float:
    return int(value) if float(value).is_integer() else value


def _registry_type(inferred_type: str) -> str:
    if inferred_type in {"number", "number_from_string", "enum"}:
        return inferred_type
    if inferred_type == "mostly_empty":
        return "string"
    return "string"


def _field_aliases(column: dict[str, Any]) -> list[str]:
    aliases = []
    for value in [column["source_column"], column["field_id"], column["label"]]:
        text = cell_text(value)
        if text and text not in aliases:
            aliases.append(text)
    return aliases


def _top_result_candidates(columns: list[dict[str, Any]]) -> list[dict[str, Any]]:
    candidates = []
    for column in columns:
        if column["pii_risk"]:
            continue
        if column["role"] in {"category", "metric", "identifier", "text"}:
            candidates.append(
                {
                    "key": column["field_id"],
                    "field_id": column["field_id"],
                    "source_column": column["source_column"],
                    "status": REVIEW_STATUS,
                }
            )
        if len(candidates) >= 8:
            break
    return candidates


def _sort_candidates(columns: list[dict[str, Any]]) -> list[dict[str, Any]]:
    candidates = []
    for column in columns:
        if column["inferred_type"] not in {"number", "number_from_string"}:
            continue
        direction = _sort_direction(column["source_column"])
        candidates.append(
            {
                "field_id": column["field_id"],
                "source_column": column["source_column"],
                "direction": direction,
                "nulls": "LAST",
                "status": REVIEW_STATUS,
            }
        )
    return candidates


def _sort_direction(source_column: Any) -> str:
    if _has_keyword(source_column, DESC_SORT_KEYWORDS):
        return "DESC"
    if _has_keyword(source_column, ASC_SORT_KEYWORDS):
        return "ASC"
    return "ASC"


def _rule_candidates(columns: list[dict[str, Any]]) -> list[dict[str, Any]]:
    candidates = []
    for column in columns:
        ops = column["candidate_allowed_ops"]
        if not ops:
            continue
        candidates.append(
            {
                "rule_id": f"candidate_{column['field_id']}",
                "category": "candidate",
                "field_id": column["field_id"],
                "source_column": column["source_column"],
                "candidate_operators": ops,
                "status": REVIEW_STATUS,
                "requires_human_confirmation": True,
                "hard_filter_allowed": False,
                "trace_reason": "自动生成候选规则，必须人工 review 后才能启用。",
            }
        )
    return candidates


def _extraction_aliases(
    columns: list[dict[str, Any]],
    llm_candidates: dict[str, Any],
) -> dict[str, Any]:
    fields = []
    for column in columns:
        fields.append(
            {
                "field_id": column["field_id"],
                "source_column": column["source_column"],
                "status": REVIEW_STATUS,
                "aliases": _field_aliases(column),
                "sample_value_alias_candidates": []
                if column["pii_risk"]
                else column["sample_values"][:5],
            }
        )
    return {
        "status": DRAFT_STATUS,
        "fields": fields,
        "llm_candidate_aliases": llm_candidates.get("candidate_aliases", []),
        "notes": [
            "自动 aliases 只能作为候选。",
            "LLM 候选 aliases 不能直接进入 RuleVerifier hard rules。",
        ],
    }


def _answer_template_seed(
    domain_id: str,
    top_result_candidates: list[dict[str, Any]],
    llm_candidates: dict[str, Any],
) -> str:
    lines = [
        f"# {domain_id} answer templates seed",
        "",
        "status: draft",
        "",
        "该文件由 schema profiling 自动生成，只能作为人工 review 起点。",
        "",
        "## 候选展示字段",
        "",
    ]
    if not top_result_candidates:
        lines.append("- 暂无候选字段。")
    for item in top_result_candidates:
        lines.append(
            f"- `{item['field_id']}` <- `{item['source_column']}` "
            f"status={item['status']}"
        )
    candidate_templates = llm_candidates.get("candidate_templates") or []
    if candidate_templates:
        lines.extend(["", "## LLM 候选模板", ""])
        for index, template in enumerate(candidate_templates, start=1):
            lines.append(f"- candidate_{index}: {template}")
    lines.extend(
        [
            "",
            "这些模板不会自动进入 `answer_templates.json`，必须人工 approve。",
        ]
    )
    return "\n".join(lines) + "\n"


def _llm_candidates(profile: dict[str, Any], llm: str) -> dict[str, Any]:
    if llm == "off":
        return {}
    from src.extractors.deepseek_extractor import DeepSeekClient

    safe_profile = _safe_llm_profile(profile)
    response = DeepSeekClient().chat_json(
        system_prompt=(
            "你是 domain pack 配置候选生成器。只返回严格 JSON。"
            "你只能基于 schema profile 和脱敏样例提出 candidate aliases "
            "或 candidate answer templates，不能输出 hard rules，不能声称可执行。"
        ),
        user_prompt=json.dumps(
            {
                "schema_profile": safe_profile,
                "required_output": {
                    "candidate_aliases": [],
                    "candidate_templates": [],
                    "safety_note": "只能作为 needs_review 候选。",
                },
            },
            ensure_ascii=False,
            indent=2,
        ),
    )
    payload = response.payload
    return {
        "candidate_aliases": payload.get("candidate_aliases") or [],
        "candidate_templates": payload.get("candidate_templates") or [],
        "usage": response.usage,
    }


def _safe_llm_profile(profile: dict[str, Any]) -> dict[str, Any]:
    columns = []
    for column in profile["columns"]:
        samples = []
        if not column.get("pii_risk"):
            samples = [
                _sanitize_llm_sample(value)
                for value in (column.get("sample_values") or [])[:3]
            ]
        columns.append(
            {
                "field_id": column["field_id"],
                "source_column": column["source_column"],
                "dtype": column["dtype"],
                "inferred_type": column["inferred_type"],
                "role": column["role"],
                "null_rate": column["null_rate"],
                "unique_count": column["unique_count"],
                "high_cardinality": column["high_cardinality"],
                "pii_risk": column["pii_risk"],
                "sample_values": samples,
                "numeric": column["numeric"],
                "keyword_flags": column["keyword_flags"],
            }
        )
    return {
        "domain_id": profile["domain_id"],
        "status": profile["status"],
        "row_count": profile["row_count"],
        "column_count": profile["column_count"],
        "columns": columns,
    }


def _sanitize_llm_sample(value: Any) -> Any:
    text = cell_text(value)
    if "@" in text:
        return "<redacted:email>"
    if re.search(r"\d{6,}", text):
        return re.sub(r"\d", "0", text[:16])
    return text[:80]


def _redact_sample(value: str) -> str:
    if "@" in value:
        return "<redacted:email>"
    if re.search(r"\d", value):
        return "<redacted:number>"
    return "<redacted>"


def _json_scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if hasattr(value, "item"):
        try:
            return value.item()
        except (TypeError, ValueError):
            pass
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _first_sortable_field(
    fields: dict[str, dict[str, Any]],
    output_field_ids: list[str],
) -> str | None:
    for field_id in output_field_ids:
        spec = fields.get(field_id) or {}
        if spec.get("type") in {"number", "number_from_string"} and (
            "sort" in spec.get("allowed_ops", [])
        ):
            return field_id
    return None


def _transform_for_type(field_type: Any) -> str:
    if field_type in {"number", "number_from_string"}:
        return "number"
    return "text"


def _money_like(source_column: Any) -> bool:
    return _has_keyword(source_column, {"price", "rent", "cost", "fee", "tuition", "价格", "租金", "费用", "学费"})


def _normalize_domain_name(domain_name: str) -> str:
    value = domain_name.strip().lower().replace("-", "_")
    if not re.fullmatch(r"[a-z][a-z0-9_]*", value):
        raise ValueError(
            "domain_name must start with a letter and contain only letters, digits, or underscores"
        )
    return value


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _to_yaml(value: Any, indent: int = 0) -> str:
    text = _yaml_lines(value, indent)
    return "\n".join(text) + "\n"


def _yaml_lines(value: Any, indent: int = 0) -> list[str]:
    pad = " " * indent
    if isinstance(value, dict):
        lines = []
        for key, item in value.items():
            if isinstance(item, (dict, list)):
                lines.append(f"{pad}{key}:")
                lines.extend(_yaml_lines(item, indent + 2))
            else:
                lines.append(f"{pad}{key}: {_yaml_scalar(item)}")
        return lines
    if isinstance(value, list):
        if not value:
            return [f"{pad}[]"]
        lines = []
        for item in value:
            if isinstance(item, dict):
                lines.append(f"{pad}-")
                lines.extend(_yaml_lines(item, indent + 2))
            elif isinstance(item, list):
                lines.append(f"{pad}-")
                lines.extend(_yaml_lines(item, indent + 2))
            else:
                lines.append(f"{pad}- {_yaml_scalar(item)}")
        return lines
    return [f"{pad}{_yaml_scalar(value)}"]


def _yaml_scalar(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    return json.dumps(str(value), ensure_ascii=False)


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Generate a draft domain pack from CSV/Excel schema profiling."
    )
    parser.add_argument("source_path", help="CSV/Excel source file.")
    parser.add_argument("domain_name", help="Domain id, e.g. housing or products.")
    parser.add_argument(
        "--output-root",
        default=str(ROOT_DIR / "domains"),
        help="Directory under which domains/<domain> will be created.",
    )
    parser.add_argument(
        "--llm",
        choices=["off", "deepseek"],
        default="off",
        help="Optional candidate config generation. Default is off.",
    )
    parser.add_argument(
        "--sheet-name",
        default=None,
        help="Excel sheet name to read. CSV input ignores this option.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = _build_arg_parser()
    args = parser.parse_args(argv)
    result = generate_domain_pack(
        source_path=args.source_path,
        domain_name=args.domain_name,
        output_root=args.output_root,
        llm=args.llm,
        sheet_name=args.sheet_name,
    )
    print(json.dumps(result.to_dict(), ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
