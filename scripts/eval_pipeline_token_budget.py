"""Pipeline-level token budget comparison.

Compares a naive "send Excel + user input to LLM" approach against the
methodology pipeline. The naive LLM path is estimated from the real workbook
serialization size; it is not executed because sending the full Excel content
would be expensive and does not provide deterministic verification.
"""

from __future__ import annotations

import json
import math
import sys
from pathlib import Path
from typing import Any

import openpyxl

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from src.adapters.excel_adapter import cell_text


WORKBOOK_NAME = "广东省2025年志愿填报大数据（24-25）0523.xlsx"
OUTPUT_DIR = Path("outputs/eval")
OUTPUT_PATH = OUTPUT_DIR / "pipeline_token_budget.json"
EVAL_MODES_PATH = OUTPUT_DIR / "eval_modes.json"

DEMO_INPUT = "我是广东物理类，排位32000，想学计算机，最好在广州深圳，学校稳一点，不想去太贵的中外合作。"
REQUIRED_COLUMNS = ["生源地", "科类", "专业名称", "城市", "专业组最低位次1", "学费"]
DEFAULT_CONTEXT_BUDGETS = [32000, 128000, 1000000]


def estimate_tokens(text: str) -> int:
    """Approximate tokens without model-specific tokenizer.

    Chinese characters are counted close to one token each; other characters are
    approximated at four characters per token. This is conservative enough for
    comparing full-Excel prompting against schema/rule pipelines.
    """

    cjk = 0
    other = 0
    for char in text:
        if "\u4e00" <= char <= "\u9fff":
            cjk += 1
        else:
            other += 1
    return cjk + math.ceil(other / 4)


def detect_header_row(sheet: Any) -> tuple[int, list[str]]:
    required = set(REQUIRED_COLUMNS)
    for row_number, row in enumerate(sheet.iter_rows(values_only=True, min_row=1, max_row=25), start=1):
        headers = [cell_text(value) for value in row]
        if required.issubset(set(headers)):
            return row_number, headers
    raise RuntimeError("无法识别表头行。")


def row_to_csv_line(values: list[Any]) -> str:
    escaped = []
    for value in values:
        text = cell_text(value).replace('"', '""')
        if "," in text or "\n" in text or '"' in text:
            text = f'"{text}"'
        escaped.append(text)
    return ",".join(escaped) + "\n"


def estimate_workbook_serialization(column_filter: list[str] | None = None) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(WORKBOOK_NAME, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header_row, headers = detect_header_row(sheet)
    selected_indices = list(range(len(headers)))
    selected_headers = headers
    if column_filter:
        header_index = {name: idx for idx, name in enumerate(headers)}
        selected_indices = [header_index[column] for column in column_filter]
        selected_headers = column_filter

    token_count = estimate_tokens(row_to_csv_line(selected_headers))
    char_count = len(row_to_csv_line(selected_headers))
    data_rows = 0
    for row in sheet.iter_rows(values_only=True, min_row=header_row + 1):
        if not any(value is not None for value in row):
            continue
        selected = [row[idx] if idx < len(row) else None for idx in selected_indices]
        line = row_to_csv_line(selected)
        token_count += estimate_tokens(line)
        char_count += len(line)
        data_rows += 1

    return {
        "sheet": sheet.title,
        "header_row": header_row,
        "columns": len(selected_headers),
        "data_rows": data_rows,
        "estimated_serialized_chars": char_count,
        "estimated_serialized_tokens": token_count,
    }


def load_methodology_results() -> dict[str, Any]:
    if not EVAL_MODES_PATH.exists():
        return {"status": "missing", "reason": "请先运行 scripts/eval_modes.py。"}
    eval_modes = json.loads(EVAL_MODES_PATH.read_text(encoding="utf-8"))
    modes = eval_modes["modes"]
    return {
        "regex_extractor_symbolic_verifier": {
            "result_count": modes["regex_extractor_symbolic_verifier"].get("result_count"),
            "task_success": modes["regex_extractor_symbolic_verifier"].get("task_success"),
            "total_tokens": modes["regex_extractor_symbolic_verifier"].get("total_tokens"),
            "trace_complete": modes["regex_extractor_symbolic_verifier"].get("trace_complete"),
        },
        "deepseek_extractor_symbolic_verifier": {
            "status": modes["deepseek_extractor_symbolic_verifier"].get("status"),
            "result_count": modes["deepseek_extractor_symbolic_verifier"].get("result_count"),
            "task_success": modes["deepseek_extractor_symbolic_verifier"].get("task_success"),
            "total_tokens": modes["deepseek_extractor_symbolic_verifier"].get("total_tokens"),
            "trace_complete": modes["deepseek_extractor_symbolic_verifier"].get("trace_complete"),
        },
        "llm_only_baseline": {
            "status": modes["llm_only_baseline"].get("status"),
            "task_success": modes["llm_only_baseline"].get("task_success"),
            "total_tokens": modes["llm_only_baseline"].get("total_tokens"),
            "unsafe_flags": modes["llm_only_baseline"].get("unsafe_flags"),
        },
        "schema_aware_llm_only_baseline": {
            "status": modes["schema_aware_llm_only_baseline"].get("status"),
            "task_success": modes["schema_aware_llm_only_baseline"].get("task_success"),
            "total_tokens": modes["schema_aware_llm_only_baseline"].get("total_tokens"),
            "unsafe_flags": modes["schema_aware_llm_only_baseline"].get("unsafe_flags"),
        },
    }


def budget_fit(estimated_tokens: int) -> dict[str, bool]:
    return {f"fits_{budget}_token_budget": estimated_tokens <= budget for budget in DEFAULT_CONTEXT_BUDGETS}


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    prompt_instruction = (
        "你是高考志愿规划助手。请阅读表格数据和用户偏好，"
        "返回推荐行和解释。\n"
    )
    prompt_tokens = estimate_tokens(prompt_instruction + DEMO_INPUT)
    full_excel = estimate_workbook_serialization()
    required_columns_excel = estimate_workbook_serialization(REQUIRED_COLUMNS)

    naive_full_tokens = prompt_tokens + full_excel["estimated_serialized_tokens"]
    naive_required_tokens = prompt_tokens + required_columns_excel["estimated_serialized_tokens"]

    result = {
        "input": DEMO_INPUT,
        "token_estimation_note": (
            "这是不依赖具体 tokenizer 的近似估算。中文字符近似按一个 token 计算；"
            "非中文字符按约四个字符一个 token 估算。"
        ),
        "comparison_goal": (
            "比较“直接把 Excel 和自然语言交给 LLM”与当前方法管线在 token 预算和可验证成功率上的差异。"
        ),
        "naive_direct_llm_full_excel": {
            "description": "把自然语言输入和完整序列化 Excel 表格一起发给 LLM，并要求它直接回答。",
            "estimated_input_tokens": naive_full_tokens,
            **budget_fit(naive_full_tokens),
            "task_success": "not_executed",
            "reason_not_executed": "完整 Excel 直接提示只做 token 预算估算，不实际调用 API。",
            "expected_weaknesses": [
                "上下文成本很高。",
                "没有确定性字段验证。",
                "容易推断不受支持的字段。",
                "不能保证逐行证据完整。",
            ],
            "workbook_serialization": full_excel,
        },
        "naive_direct_llm_required_columns": {
            "description": "只发送 MVP 必需列的全部行，再加自然语言输入。",
            "estimated_input_tokens": naive_required_tokens,
            **budget_fit(naive_required_tokens),
            "task_success": "not_executed",
            "reason_not_executed": "仍然没有确定性验证器，只作为直接提示的下界估算。",
            "workbook_serialization": required_columns_excel,
        },
        "methodology_pipeline": load_methodology_results(),
        "main_claim": (
            "对于明确需求，符号管线在抽取阶段消耗零或少量 LLM token，随后执行带 trace 的确定性筛选。"
            "直接把 Excel 和 prompt 交给 LLM 会消耗数量级更高的 token，仍不能保证字段接地或不可执行偏好的拒绝。"
        ),
    }
    OUTPUT_PATH.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}")
    print("naive full Excel estimated tokens:", naive_full_tokens)
    print("naive required columns estimated tokens:", naive_required_tokens)
    methodology = result["methodology_pipeline"]
    if "regex_extractor_symbolic_verifier" in methodology:
        print("regex methodology tokens:", methodology["regex_extractor_symbolic_verifier"]["total_tokens"])
        print("regex methodology result_count:", methodology["regex_extractor_symbolic_verifier"]["result_count"])
    if "deepseek_extractor_symbolic_verifier" in methodology:
        print("deepseek methodology tokens:", methodology["deepseek_extractor_symbolic_verifier"]["total_tokens"])
        print("deepseek methodology result_count:", methodology["deepseek_extractor_symbolic_verifier"]["result_count"])


if __name__ == "__main__":
    main()
