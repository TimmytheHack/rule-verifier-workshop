"""Excel data adapter for the first MVP backend."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd


def cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


@dataclass(frozen=True)
class ExcelDataSet:
    """Excel 数据源的最小结构化表示。"""

    workbook_path: Path
    sheet_name: str
    header_row: int
    headers: list[str]
    header_index: dict[str, int]
    dataframe: pd.DataFrame


class ExcelAdapter:
    """读取 Excel，并检测真实表头行。

    Adapter 只报告数据实际包含什么，不解释用户偏好。
    """

    def __init__(self, workbook_path: str | Path, required_columns: list[str]) -> None:
        self.workbook_path = Path(workbook_path)
        self.required_columns = required_columns

    def load(self) -> ExcelDataSet:
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.workbook_path}")

        workbook = openpyxl.load_workbook(self.workbook_path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        header_row, headers = self._detect_header_row(sheet)
        dataframe = pd.read_excel(
            self.workbook_path,
            sheet_name=sheet.title,
            header=header_row - 1,
            dtype=object,
            engine="openpyxl",
        )
        dataframe = dataframe.dropna(how="all")
        dataframe.columns = [cell_text(column) for column in dataframe.columns]

        header_index = {name: idx for idx, name in enumerate(headers) if name}
        return ExcelDataSet(
            workbook_path=self.workbook_path,
            sheet_name=sheet.title,
            header_row=header_row,
            headers=headers,
            header_index=header_index,
            dataframe=dataframe,
        )

    def _detect_header_row(self, sheet: Any) -> tuple[int, list[str]]:
        required = set(self.required_columns)
        for row_number, row in enumerate(
            sheet.iter_rows(values_only=True, min_row=1, max_row=25),
            start=1,
        ):
            headers = [cell_text(value) for value in row]
            if required.issubset(set(headers)):
                return row_number, headers
        raise RuntimeError(f"Could not find a header row containing: {', '.join(self.required_columns)}")
